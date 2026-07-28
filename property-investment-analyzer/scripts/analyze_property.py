#!/usr/bin/env python3
"""Rental / income-property investment analyzer.

Underwrites one or many buy-and-hold rental opportunities and builds an Excel
workbook: a ranked SCREENING matrix + a full DETAIL underwriting tab per
property + shared ASSUMPTIONS. Every metric is a live formula.

Usage:
    python3 analyze_property.py --out deals.xlsx
    python3 analyze_property.py --config deals.json --out deals.xlsx
    # then embed values so it opens populated + zero-error check:
    python3 validate_and_fill.py deals.xlsx

config JSON shape:
  {"assumptions": {...overrides...},
   "properties": [
     {"name","price","units","sqft","rent_actual","rent_market",
      "taxes","insurance","hoa_mo","rehab"}, ...]}
Defaults = realistic Marion County FL rentals.
"""
import argparse, json, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

FONT="Arial"
NAVY="1F3A5F"; TEAL="143642"; LGREY="EEF1F4"; PAPER="FAF8F4"
BLUE="0000FF"; BLACK="000000"; GREEN="008000"; WHITE="FFFFFF"; YEL="FFF3C4"
GOODBG="D6E4D0"; WARNBG="F4EBD6"; BADBG="F2E0DC"
CUR='$#,##0;($#,##0);"-"'; CUR2='$#,##0.00;($#,##0.00);"-"'
PCT='0.0%'; PCT2='0.00%'; MULT='0.00"x"'; NUM='#,##0;(#,##0);"-"'
thin=Side(style="thin",color="B8BFC7"); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)

DEFAULTS = {
  "assumptions": {
    "vacancy":0.06, "mgmt":0.10, "maintenance":0.05, "reserve_unit":350,
    "other_opex":600, "down":0.25, "rate":0.0675, "amort":30, "closing":0.03,
    "rent_growth":0.03, "exp_growth":0.03, "apprec":0.03, "exit_cap":0.07, "hold_years":5,
    "sell_cost":0.07, "target_cap":0.07, "target_coc":0.08, "target_dscr":1.20,
  },
  "properties": [
    {"name":"Marion Oaks 3/2 SFR","price":265000,"units":1,"sqft":1500,
     "rent_actual":1800,"rent_market":1900,"taxes":2900,"insurance":1750,"hoa_mo":0,"rehab":0},
    {"name":"Silver Springs Shores 4/2","price":245000,"units":1,"sqft":1751,
     "rent_actual":1950,"rent_market":2050,"taxes":2700,"insurance":1750,"hoa_mo":0,"rehab":5000},
    {"name":"Marion Oaks duplex 2x2","price":399900,"units":2,"sqft":1968,
     "rent_actual":2600,"rent_market":2900,"taxes":4200,"insurance":3000,"hoa_mo":0,"rehab":0},
  ],
}

def C(ws,coord,val=None,*,bold=False,italic=False,size=10,color=BLACK,fill=None,
      fmt=None,align=None,wrap=False,border=False,valign=None):
    c=ws[coord]
    if val is not None: c.value=val
    c.font=Font(name=FONT,bold=bold,italic=italic,size=size,color=color)
    if fill: c.fill=PatternFill("solid",fgColor=fill)
    if fmt: c.number_format=fmt
    if align or wrap or valign: c.alignment=Alignment(horizontal=align,wrap_text=wrap,vertical=valign or "center")
    if border: c.border=BORD
    return c
def band(ws,row,text,sub=None,span=8):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=span)
    C(ws,f"A{row}",text,bold=True,size=15,color=WHITE,fill=TEAL); ws.row_dimensions[row].height=26
    if sub:
        ws.merge_cells(start_row=row+1,start_column=1,end_row=row+1,end_column=span)
        C(ws,f"A{row+1}",sub,italic=True,size=9,color="5A6470",fill=PAPER); ws.row_dimensions[row+1].height=16
        return row+2
    return row+1
def sect(ws,row,text,span=8,fill=NAVY):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=span)
    C(ws,f"A{row}",text,bold=True,size=11,color=WHITE,fill=fill); ws.row_dimensions[row].height=20

# ---- fixed DETAIL layout rows (so SCREENING can reference them) ----
R = dict(price=4,units=5,sqft=6,ra=7,rm=8,taxes=9,ins=10,hoa=11,rehab=12,
         colhdr=13,gsr=14,vac=15,egi=16,mgmt=17,maint=18,tx=19,insr=20,res=21,
         hoar=22,oth=23,opx=24,noi=25,
         basis=27,capip=28,cappf=29,grm=30,onepct=31,ppu=32,psf=33,exr=34,
         down=36,loan=37,close=38,ctc=39,mds=40,ads=41,dscrip=42,dscrpf=43,beocc=44,
         cfip=46,cocip=47,cfpf=48,cocpf=49,
         noilift=51,valcreate=52,
         moxcap=54,mox1=55,moxdscr=56,
         yhdr=58,ecf=59,irr=61,em=62,verdict=64)

def build_detail(ws, p, S):
    ws.sheet_view.showGridLines=False
    for col,w in zip("ABCD",[34,16,16,30]): ws.column_dimensions[col].width=w
    band(ws,1,p["name"],"Buy-and-hold rental underwriting  ·  In-place vs Pro-forma",span=4)
    def inp(row,label,val,fmt=CUR,note=""):
        C(ws,f"A{row}",label); C(ws,f"B{row}",val,color=BLUE,fmt=fmt,align="right")
        if note: C(ws,f"C{row}",note,italic=True,size=8,color="6A7480")
    sect(ws,3,"INPUTS  (blue = edit)",span=4,fill=NAVY)
    inp(R["price"],"Purchase price",p["price"])
    inp(R["units"],"Units",p["units"],NUM)
    inp(R["sqft"],"Building SF",p["sqft"],NUM)
    inp(R["ra"],"Actual/current rent (total /mo)",p["rent_actual"])
    inp(R["rm"],"Market/pro-forma rent (total /mo)",p["rent_market"])
    inp(R["taxes"],"Property taxes /yr",p["taxes"])
    inp(R["ins"],"Insurance /yr",p["insurance"])
    inp(R["hoa"],"HOA /mo",p.get("hoa_mo",0))
    inp(R["rehab"],"Rehab / capex at purchase",p.get("rehab",0))
    b=lambda k:f"B{R[k]}"; c=lambda k:f"C{R[k]}"
    def frow(row,label,bexpr,cexpr=None,fmt=CUR,bold=False,fill=None,pct=False,mult=False,note=""):
        f=PCT if pct else (MULT if mult else fmt)
        C(ws,f"A{row}",label,bold=bold,fill=fill)
        C(ws,f"B{row}",bexpr,color=BLACK,fmt=f,align="right",bold=bold,fill=fill)
        if cexpr is not None: C(ws,f"C{row}",cexpr,color=BLACK,fmt=f,align="right",bold=bold,fill=fill)
        if note: C(ws,f"D{row}",note,italic=True,size=8,color="6A7480")
    C(ws,f"B{R['colhdr']}","In-place",bold=True,align="right",fill=LGREY)
    C(ws,f"C{R['colhdr']}","Pro-forma",bold=True,align="right",fill=LGREY)
    C(ws,f"A{R['colhdr']}","OPERATING STATEMENT",bold=True,color=WHITE,fill=NAVY)
    frow(R["gsr"],"Gross scheduled rent",f"={b('ra')}*12",f"={b('rm')}*12")
    frow(R["vac"],"Vacancy & credit loss",f"=-{b('gsr')}*{S['vacancy']}",f"=-{c('gsr')}*{S['vacancy']}")
    frow(R["egi"],"Effective gross income",f"={b('gsr')}+{b('vac')}",f"={c('gsr')}+{c('vac')}",bold=True)
    frow(R["mgmt"],"Property management",f"=-{b('egi')}*{S['mgmt']}",f"=-{c('egi')}*{S['mgmt']}")
    frow(R["maint"],"Repairs & maintenance",f"=-{b('egi')}*{S['maintenance']}",f"=-{c('egi')}*{S['maintenance']}")
    frow(R["tx"],"Property taxes",f"=-{b('taxes')}",f"=-{b('taxes')}")
    frow(R["insr"],"Insurance",f"=-{b('ins')}",f"=-{b('ins')}")
    frow(R["res"],"Capital reserve",f"=-{S['reserve_unit']}*{b('units')}",f"=-{S['reserve_unit']}*{b('units')}")
    frow(R["hoar"],"HOA",f"=-{b('hoa')}*12",f"=-{b('hoa')}*12")
    frow(R["oth"],"Other (admin/legal)",f"=-{S['other_opex']}",f"=-{S['other_opex']}")
    frow(R["opx"],"Total operating expenses",f"=SUM(B{R['mgmt']}:B{R['oth']})",f"=SUM(C{R['mgmt']}:C{R['oth']})",bold=True)
    frow(R["noi"],"NET OPERATING INCOME",f"={b('egi')}+{b('opx')}",f"={c('egi')}+{c('opx')}",bold=True,fill=GOODBG)
    sect(ws,26,"RETURNS (unlevered)",span=4,fill=NAVY)
    frow(R["basis"],"All-in basis (price + rehab)",f"={b('price')}+{b('rehab')}",bold=True)
    frow(R["capip"],"Cap rate — in-place",f"={b('noi')}/{b('basis')}",pct=True,bold=True,fill=YEL)
    frow(R["cappf"],"Cap rate — pro-forma",f"={c('noi')}/{b('basis')}",pct=True,bold=True)
    frow(R["grm"],"Gross rent multiplier",f"={b('price')}/{b('gsr')}",fmt=MULT)
    frow(R["onepct"],"1% rule (rent/price)",f"={b('ra')}/{b('price')}",pct=True,note="target ≥1.0%")
    frow(R["ppu"],"Price per unit",f"={b('price')}/{b('units')}")
    frow(R["psf"],"Price per SF",f"={b('price')}/{b('sqft')}",fmt=CUR2)
    frow(R["exr"],"Operating expense ratio",f"=-{b('opx')}/{b('egi')}",pct=True)
    sect(ws,35,"FINANCING",span=4,fill=NAVY)
    frow(R["down"],"Down payment",f"={b('price')}*{S['down']}")
    frow(R["loan"],"Loan amount",f"={b('price')}-{b('down')}")
    frow(R["close"],"Closing costs",f"={b('price')}*{S['closing']}")
    frow(R["ctc"],"CASH TO CLOSE (incl rehab)",f"={b('down')}+{b('close')}+{b('rehab')}",bold=True,fill=YEL)
    frow(R["mds"],"Monthly debt service",f"=PMT({S['rate']}/12,{S['amort']}*12,-{b('loan')})")
    frow(R["ads"],"Annual debt service",f"={b('mds')}*12")
    frow(R["dscrip"],"DSCR — in-place",f"={b('noi')}/{b('ads')}",mult=True,bold=True,fill=YEL,note="lender ≥1.20-1.25")
    frow(R["dscrpf"],"DSCR — pro-forma",f"={c('noi')}/{b('ads')}",mult=True)
    frow(R["beocc"],"Breakeven occupancy",f"=(-{b('opx')}+{b('ads')})/{b('gsr')}",pct=True)
    sect(ws,45,"LEVERED CASH FLOW",span=4,fill=NAVY)
    frow(R["cfip"],"Cash flow before tax — in-place",f"={b('noi')}-{b('ads')}",bold=True)
    frow(R["cocip"],"Cash-on-cash — in-place",f"={b('cfip')}/{b('ctc')}",pct=True,bold=True,fill=YEL,note="target ≥8%")
    frow(R["cfpf"],"Cash flow before tax — pro-forma",f"={c('noi')}-{b('ads')}",bold=True)
    frow(R["cocpf"],"Cash-on-cash — pro-forma",f"={b('cfpf')}/{b('ctc')}",pct=True,bold=True)
    sect(ws,50,"VALUE-ADD",span=4,fill=NAVY)
    frow(R["noilift"],"NOI lift (pro-forma − in-place)",f"={c('noi')}-{b('noi')}",bold=True)
    frow(R["valcreate"],"Value created (lift ÷ exit cap)",f"={b('noilift')}/{S['exit_cap']}",bold=True,fill=GOODBG)
    sect(ws,53,"MAX SUPPORTABLE OFFER",span=4,fill=NAVY)
    frow(R["moxcap"],f"@ target cap ({S['target_cap']*100:.1f}%)",f"={b('noi')}/{S['target_cap']}",bold=True)
    frow(R["mox1"],"@ 1% rule",f"={b('ra')}*100")
    frow(R["moxdscr"],f"@ target DSCR ({S['target_dscr']}x)",
         f"=(-PV({S['rate']}/12,{S['amort']}*12,({b('noi')}/{S['target_dscr']})/12))/(1-{S['down']})",bold=True)
    sect(ws,57,f"{S['hold_years']}-YEAR HOLD (pro-forma basis)",span=8,fill=TEAL)
    hy=int(S["hold_years"])
    C(ws,f"A{R['yhdr']}","Year",bold=True,fill=LGREY)
    for y in range(0,hy+1): C(ws,get_column_letter(2+y)+str(R['yhdr']),y,bold=True,fill=LGREY,align="right",fmt=NUM)
    C(ws,f"A{R['ecf']}","Investor equity CF",bold=True)
    C(ws,f"B{R['ecf']}",f"=-{b('ctc')}",fmt=CUR,align="right")
    for y in range(1,hy+1):
        col=get_column_letter(2+y)
        cf=f"{c('noi')}*(1+{S['rent_growth']})^{y-1}-{b('ads')}"
        if y==hy:
            bal=f"MAX(0,{b('loan')}*(1+{S['rate']}/12)^({hy}*12)-{b('mds')}*(((1+{S['rate']}/12)^({hy}*12)-1)/({S['rate']}/12)))"
            sale=f"({b('price')}*(1+{S['apprec']})^{hy})*(1-{S['sell_cost']})"  # appreciation-based exit
            expr=f"={cf}+{sale}-{bal}"
        else:
            expr=f"={cf}"
        C(ws,f"{col}{R['ecf']}",expr,fmt=CUR,align="right")
    last=get_column_letter(2+hy)
    frow(R["irr"],f"{hy}-yr levered IRR",f"=IRR(B{R['ecf']}:{last}{R['ecf']})",pct=True,bold=True,fill=YEL)
    frow(R["em"],"Equity multiple",f"=SUM(C{R['ecf']}:{last}{R['ecf']})/(-B{R['ecf']})",mult=True,bold=True)
    sect(ws,63,"VERDICT",span=4,fill=TEAL)
    C(ws,f"A{R['verdict']}","Screen result",bold=True)
    C(ws,f"B{R['verdict']}",
      f'=IF(AND({b("capip")}>={S["target_cap"]},{b("dscrip")}>={S["target_dscr"]},{b("cocip")}>={S["target_coc"]}),"BUY",'
      f'IF(OR({b("cappf")}>={S["target_cap"]},{b("cocpf")}>={S["target_coc"]}),"WATCH","PASS"))',
      bold=True,align="center")
    ws.freeze_panes="A4"

def main():
    ap=argparse.ArgumentParser(description=__doc__,formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--config"); ap.add_argument("--out",default="rental-deals.xlsx")
    a=ap.parse_args()
    cfg=json.loads(json.dumps(DEFAULTS))
    if a.config:
        u=json.load(open(a.config))
        if "assumptions" in u: cfg["assumptions"].update(u["assumptions"])
        if "properties" in u: cfg["properties"]=u["properties"]
    S=cfg["assumptions"]; props=cfg["properties"]
    wb=Workbook()

    # ASSUMPTIONS
    aw=wb.active; aw.title="Assumptions"; aw.sheet_view.showGridLines=False
    for col,w in zip("ABCD",[34,14,10,36]): aw.column_dimensions[col].width=w
    r=band(aw,1,"UNDERWRITING ASSUMPTIONS","Shared across all deals. Edit the blue cells in each Detail tab for per-property inputs.",span=4)
    r+=2
    rows=[("Vacancy & credit loss","vacancy",PCT),("Property management % EGI","mgmt",PCT),
          ("Repairs & maintenance % EGI","maintenance",PCT),("Capital reserve / unit / yr","reserve_unit",CUR),
          ("Other opex / yr","other_opex",CUR),("Down payment %","down",PCT),("Loan rate","rate",PCT),
          ("Amortization (yrs)","amort",NUM),("Closing costs %","closing",PCT),("Rent growth /yr","rent_growth",PCT),
          ("Expense growth /yr","exp_growth",PCT),("Appreciation /yr","apprec",PCT),
          ("Exit cap (value-add only)","exit_cap",PCT),("Hold period (yrs)","hold_years",NUM),
          ("Selling cost %","sell_cost",PCT),("Target cap (screen)","target_cap",PCT),
          ("Target cash-on-cash","target_coc",PCT),("Target DSCR","target_dscr",MULT)]
    for label,key,fmt in rows:
        C(aw,f"A{r}",label); C(aw,f"B{r}",S[key],color=BLUE,fmt=fmt,align="right",fill=YEL if key in("target_cap","target_coc","target_dscr") else None)
        r+=1
    aw.freeze_panes="A3"

    # DETAIL tabs
    def safe(t): return re.sub(r'[\\/?*\[\]:]', '-', t)
    detail=[]   # (sheet_title, display_name)
    for i,p in enumerate(props,1):
        nm=safe(f"D{i} {p['name']}")[:31]
        ws=wb.create_sheet(nm); detail.append((ws.title,p["name"]))
        build_detail(ws,p,S)
    detail_sheets=[t for t,_ in detail]

    # SCREENING
    sc=wb.create_sheet("Screening"); sc.sheet_view.showGridLines=False
    cols=[("Rank",8),("Property",26),("Price",13),("$/unit",12),("1% rule",9),
          ("Cap in-place",12),("Cap pro-forma",13),("DSCR",9),("Cash-on-cash",13),
          (f"{int(S['hold_years'])}-yr IRR",10),("Max @ cap",13),("Score",9),("Verdict",11)]
    for j,(_,w) in enumerate(cols): sc.column_dimensions[get_column_letter(1+j)].width=w
    r=band(sc,1,"OPPORTUNITY SCREENING MATRIX","Ranked by blended yield score. Green = Buy · Amber = Watch · Red = Pass.",span=len(cols))
    r+=1
    hr=r
    for j,(name,_) in enumerate(cols):
        C(sc,get_column_letter(1+j)+str(r),name,bold=True,color=WHITE,fill=NAVY,align="center" if j else "left",border=True)
    r+=1
    first=r
    def dref(sh,key): return f"'{sh}'!$B${R[key]}"
    for sh,dispname in detail:
        C(sc,f"A{r}","",align="center",border=True)  # rank filled below
        C(sc,f"B{r}",dispname,border=True)
        C(sc,f"C{r}",f"={dref(sh,'price')}",fmt=CUR,align="right",border=True,color=GREEN)
        C(sc,f"D{r}",f"={dref(sh,'ppu')}",fmt=CUR,align="right",border=True,color=GREEN)
        C(sc,f"E{r}",f"={dref(sh,'onepct')}",fmt=PCT,align="right",border=True,color=GREEN)
        C(sc,f"F{r}",f"={dref(sh,'capip')}",fmt=PCT,align="right",border=True,color=GREEN)
        C(sc,f"G{r}",f"={dref(sh,'cappf')}",fmt=PCT,align="right",border=True,color=GREEN)
        C(sc,f"H{r}",f"={dref(sh,'dscrip')}",fmt=MULT,align="right",border=True,color=GREEN)
        C(sc,f"I{r}",f"={dref(sh,'cocip')}",fmt=PCT,align="right",border=True,color=GREEN)
        C(sc,f"J{r}",f"={dref(sh,'irr')}",fmt=PCT,align="right",border=True,color=GREEN)
        C(sc,f"K{r}",f"={dref(sh,'moxcap')}",fmt=CUR,align="right",border=True,color=GREEN)
        # score = cap in-place*100 + coc*100 + dscr cushion*5
        C(sc,f"L{r}",f"=F{r}*100+I{r}*100+MAX(0,H{r}-1.2)*5",fmt='0.0',align="right",border=True)
        C(sc,f"M{r}",f"={dref(sh,'verdict')}",align="center",border=True,bold=True)
        r+=1
    last=r-1
    for rr in range(first,last+1):
        sc[f"A{rr}"]=f"=RANK(L{rr},$L${first}:$L${last},0)"; sc[f"A{rr}"].number_format="0"
        sc[f"A{rr}"].alignment=Alignment(horizontal="center"); sc[f"A{rr}"].font=Font(name=FONT,bold=True); sc[f"A{rr}"].border=BORD
    # conditional formatting on verdict column M
    rng=f"M{first}:M{last}"
    sc.conditional_formatting.add(rng,FormulaRule(formula=[f'EXACT(M{first},"BUY")'],fill=PatternFill("solid",fgColor=GOODBG)))
    sc.conditional_formatting.add(rng,FormulaRule(formula=[f'EXACT(M{first},"WATCH")'],fill=PatternFill("solid",fgColor=WARNBG)))
    sc.conditional_formatting.add(rng,FormulaRule(formula=[f'EXACT(M{first},"PASS")'],fill=PatternFill("solid",fgColor=BADBG)))
    C(sc,f"A{last+2}","Metrics per deal are on the Detail tabs. Change any blue input and everything recalculates.",italic=True,size=9,color="6A7480")
    sc.freeze_panes="A4"

    order=["Screening","Assumptions"]+detail_sheets
    for i,nm in enumerate(order): wb.move_sheet(nm,-(wb.sheetnames.index(nm))+i)
    wb.save(a.out)
    print("Saved:",a.out); print("Tabs:",wb.sheetnames)
    print("NEXT: python3 validate_and_fill.py",a.out)

if __name__=="__main__":
    main()

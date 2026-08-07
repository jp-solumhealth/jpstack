#!/usr/bin/env python3
"""validate_model.py — sanity-check a DCF .xlsx built by dcf.py.

Reads the hidden _meta sheet (raw inputs as JSON), independently RE-DERIVES the base-case
DCF in Python, and checks the hard rules an analyst would check by hand:
  1. terminal growth g < WACC            (CRITICAL — model is invalid otherwise)
  2. WACC within a sane band (5%–20%)    (warn)
  3. terminal value 40%–80% of EV        (warn — >80% means the explicit forecast is too short)
Also scans every cell of every sheet for Excel error strings (#REF!, #DIV/0!, etc.).

Usage:  python3 validate_model.py NVDA_dcf_2026-06-03.xlsx
Exit code 0 if no CRITICAL failure, 1 otherwise.
"""
import argparse, json, sys
from openpyxl import load_workbook

ERRORS = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NULL!", "#NUM!", "#N/A"}


def wacc_of(d):
    E = d["price"] * d["shares"]
    D = max(d["net_debt"], 0)
    ke = d["rf"] + d["beta"] * d["erp"]
    kd = d["cost_debt_pretax"] * (1 - d["tax_rate"])
    return E / (E + D) * ke + D / (E + D) * kd


def recompute_base(d):
    n = int(d["years"])
    w = wacc_of(d)
    g = d["terminal_growth"]
    cagr = d["scenarios"]["base"]["cagr"]
    tgt = d["scenarios"]["base"]["target_margin"]
    cur = d["current_ebit_margin"]
    rev_prev = d["revenue_base"]
    pv_sum, ufcf_last = 0.0, 0.0
    for t in range(1, n + 1):
        rev = rev_prev * (1 + cagr)
        margin = cur + (tgt - cur) * t / n
        ebit = rev * margin
        nopat = ebit * (1 - d["tax_rate"])
        da = rev * d["da_pct"]
        capex = rev * d["capex_pct"]
        dnwc = (rev - rev_prev) * d["nwc_pct"]
        ufcf = nopat + da - capex - dnwc
        period = t - 0.5
        pv_sum += ufcf / (1 + w) ** period
        ufcf_last = ufcf
        rev_prev = rev
    tv = ufcf_last * (1 + g) / (w - g) if w > g else float("inf")
    pv_tv = tv / (1 + w) ** (n - 0.5)
    ev = pv_sum + pv_tv
    equity = ev - d["net_debt"]
    return {"wacc": w, "g": g, "ev": ev, "pv_tv": pv_tv,
            "tv_pct": (pv_tv / ev if ev else float("inf")),
            "equity": equity, "per_share": equity / d["shares"]}


def main():
    ap = argparse.ArgumentParser(description="Validate a DCF .xlsx from dcf.py.")
    ap.add_argument("xlsx")
    args = ap.parse_args()

    wb = load_workbook(args.xlsx, data_only=False)
    if "_meta" not in wb.sheetnames:
        print("FAIL: no _meta sheet — was this built by dcf.py?")
        sys.exit(1)
    d = json.loads(wb["_meta"]["A1"].value)

    crit, warn, ok = [], [], []
    r = recompute_base(d)

    # 1. g < WACC
    if r["g"] < r["wacc"]:
        ok.append(f"terminal g {r['g']:.2%} < WACC {r['wacc']:.2%}")
    else:
        crit.append(f"terminal g {r['g']:.2%} >= WACC {r['wacc']:.2%} — model invalid")
    # 2. WACC band
    if 0.05 <= r["wacc"] <= 0.20:
        ok.append(f"WACC {r['wacc']:.2%} in 5–20% band")
    else:
        warn.append(f"WACC {r['wacc']:.2%} outside 5–20% — re-check β/Rf/ERP/weights")
    # 3. TV % of EV
    if 0.40 <= r["tv_pct"] <= 0.80:
        ok.append(f"TV is {r['tv_pct']:.0%} of EV (40–80% target)")
    elif r["tv_pct"] > 0.80:
        warn.append(f"TV is {r['tv_pct']:.0%} of EV (>80%) — forecast too short / g too high")
    else:
        warn.append(f"TV is {r['tv_pct']:.0%} of EV (<40%) — unusually low; check terminal assumptions")

    # 4. error-cell scan (only sees cached values if the file was opened by Excel/LibreOffice)
    wb2 = load_workbook(args.xlsx, data_only=True)
    bad = []
    for ws in wb2.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.strip() in ERRORS:
                    bad.append(f"{ws.title}!{c.coordinate}={c.value}")
    if bad:
        crit.append("error cells: " + ", ".join(bad[:10]) + (" …" if len(bad) > 10 else ""))
    else:
        ok.append("no Excel error cells found")

    print(f"== DCF validation: {args.xlsx} ({d['ticker']}) ==")
    print(f"   base value/share ≈ ${r['per_share']:.2f}  |  EV ≈ {r['ev']:,.0f}M  |  WACC {r['wacc']:.2%}")
    for m in ok:
        print(f"  [ok]   {m}")
    for m in warn:
        print(f"  [warn] {m}")
    for m in crit:
        print(f"  [FAIL] {m}")
    if crit:
        print("\nRESULT: FAIL — fix the critical issue(s) and rebuild before trusting/citing this model.")
        sys.exit(1)
    print("\nRESULT: PASS (review warnings). Reconcile with comps before any rating.")


if __name__ == "__main__":
    main()

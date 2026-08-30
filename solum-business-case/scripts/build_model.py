#!/usr/bin/env python3
"""Build the Solum business-case Excel model from a config.json.
Usage: python3 build_model.py <config.json>
Produces <output_dir>/<Client>_BusinessCase.xlsx with 4 tabs.
All results are Excel formulas (editable). See references/financial-model.md.
"""
import json, sys, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import BarChart, Reference

NAVY="011C40"; BLUE="468AF7"; TEAL="70D3C6"; ALT="F2F4F9"; BG="F2F2F9"; WHITE="FFFFFF"
INK="0F1B2D"; INP="0000FF"; MUT="6C7689"; GREY="8795AD"; HDR="33405E"; F="DM Sans"
def fnt(s=11,b=False,c=INK,i=False): return Font(name=F,size=s,bold=b,color=c,italic=i)
def fl(c): return PatternFill("solid",start_color=c,end_color=c)
TH=Side(style="thin",color="DCE2EC"); BD=Border(TH,TH,TH,TH)
CTR=Alignment(horizontal="center",vertical="center"); RGT=Alignment(horizontal="right",vertical="center")
LFT=Alignment(horizontal="left",vertical="center",wrap_text=True)
USD="$#,##0"; USDc="$#,##0.00"; MULT='0.0"×"'; NUM="#,##0"; HRS='#,##0 "hrs"'; MO='0.0 "mo"'
SKILL_DIR=os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

def fillrow(ws,r,c1,c2,color):
    for c in range(c1,c2+1): ws.cell(r,c).fill=fl(color)
def band(ws,title,sub,lastc,right=False):
    n=ord(lastc)-64
    ws.merge_cells(f"A1:{lastc}1"); ws.merge_cells(f"A2:{lastc}2"); ws.merge_cells(f"A3:{lastc}3")
    al=Alignment(horizontal="right" if right else "left",vertical="center",indent=1)
    ws["A1"]=title; ws["A1"].font=fnt(18,True,WHITE); ws["A1"].alignment=al
    ws["A2"]=sub; ws["A2"].font=fnt(10,False,"AEBBD0"); ws["A2"].alignment=al
    fillrow(ws,1,1,n,NAVY); fillrow(ws,2,1,n,NAVY); fillrow(ws,3,1,n,BLUE)
    ws.row_dimensions[1].height=30; ws.row_dimensions[2].height=17; ws.row_dimensions[3].height=4
def sec(ws,r,t,lastc):
    n=ord(lastc)-64; ws.merge_cells(f"A{r}:{lastc}{r}"); c=ws[f"A{r}"]; c.value=t
    c.font=fnt(10.5,True,WHITE); c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    fillrow(ws,r,1,n,NAVY); ws.row_dimensions[r].height=19

def main(cfg_path):
    cfg=json.load(open(cfg_path))
    client=cfg["client"]; date=cfg.get("date",""); ramp=cfg.get("ramp_months",1)
    roi_mode=cfg.get("roi_mode","recurring")
    pricing=cfg["pricing"]; one_time=cfg["one_time"]; drivers=cfg["drivers"]
    outdir=cfg["output_dir"]; os.makedirs(outdir,exist_ok=True)
    wb=Workbook()

    # ============ FINANCIAL IMPACT (engine) ============
    fi=wb.create_sheet("Financial Impact"); fi.sheet_view.showGridLines=False
    for col,w in zip("ABCDEF",[36,12,11,13,13,13]): fi.column_dimensions[col].width=w
    band(fi,"Financial Impact — Variable Model","Edit blue cells · Conservative / Expected / Best paths recalculate automatically","F")
    sec(fi,5,"RECURRING PRICING (per month)","F")
    for i,h in enumerate(["Service","Volume","Rate","Monthly","",""]):
        c=fi.cell(6,i+1,h); c.font=fnt(9.5,True,WHITE); c.fill=fl(HDR); c.alignment=CTR if i in(1,2,3) else LFT; c.border=BD
    r=7
    for i,p in enumerate(pricing):
        rr=r+i; a=ALT if i%2 else WHITE
        fi.cell(rr,1,p["service"]).font=fnt(10); fi.cell(rr,1).alignment=LFT
        cv=fi.cell(rr,2,p["volume"]); cv.font=fnt(10,True,INP); cv.alignment=CTR; cv.number_format=NUM
        cr=fi.cell(rr,3,p["rate"]); cr.font=fnt(10,True,INP); cr.alignment=CTR; cr.number_format=USDc
        cm=fi.cell(rr,4,f"=B{rr}*C{rr}"); cm.font=fnt(10); cm.alignment=RGT; cm.number_format=USD
        for c0 in range(1,5):
            fi.cell(rr,c0).border=BD
            if c0 in(1,4): fi.cell(rr,c0).fill=fl(a)
    gt=r+len(pricing)
    fi.cell(gt,1,"GRAND TOTAL — Monthly").font=fnt(11,True,WHITE)
    fi.cell(gt,4,f"=SUM(D{r}:D{gt-1})").font=fnt(11,True,WHITE); fi.cell(gt,4).alignment=RGT; fi.cell(gt,4).number_format=USD
    for c0 in range(1,5): fi.cell(gt,c0).fill=fl(NAVY); fi.cell(gt,c0).border=BD
    ann=gt+1
    fi.cell(ann,1,"Annual recurring (×12)").font=fnt(10,True,NAVY)
    fi.cell(ann,4,f"=D{gt}*12").font=fnt(10,True,NAVY); fi.cell(ann,4).alignment=RGT; fi.cell(ann,4).number_format=USD
    GT=f"D{gt}"; ANN=f"D{ann}"
    # one-time
    ot=ann+2; sec(fi,ot,"ONE-TIME (setup & integration)","F"); orow=ot+1
    for i,o in enumerate(one_time):
        rr=orow+i
        fi.cell(rr,1,o["label"]).font=fnt(10)
        cv=fi.cell(rr,4,o["amount"]); cv.font=fnt(10,True,INP); cv.alignment=RGT; cv.number_format=USD
        for c0 in range(1,5): fi.cell(rr,c0).border=BD
    if cfg.get("one_time_note"): fi.cell(orow,5,cfg["one_time_note"]).font=fnt(9,False,MUT,True)
    ott=orow+len(one_time)
    fi.cell(ott,1,"One-time total").font=fnt(10,True,NAVY)
    fi.cell(ott,4,f"=SUM(D{orow}:D{ott-1})").font=fnt(10,True,NAVY); fi.cell(ott,4).alignment=RGT; fi.cell(ott,4).number_format=USD
    y1c=ott+1
    fi.cell(y1c,1,"Total Year-1 investment (platform + setup + integration)").font=fnt(10,True,NAVY)
    fi.cell(y1c,4,f"={ANN}+D{ott}").font=fnt(10,True,NAVY); fi.cell(y1c,4).alignment=RGT; fi.cell(y1c,4).number_format=USD
    rampr=y1c+1
    fi.cell(rampr,1,"Implementation ramp (value starts at go-live)").font=fnt(10)
    cv=fi.cell(rampr,4,ramp); cv.font=fnt(10,True,INP); cv.alignment=RGT; cv.number_format='0 "mo"'
    for rr in range(ott,rampr+1):
        for c0 in range(1,5): fi.cell(rr,c0).border=BD
    OTT=f"D{ott}"; Y1C=f"D{y1c}"; RAMP=f"D{rampr}"
    # value drivers
    vs=rampr+2; sec(fi,vs,"VALUE DRIVERS  (blue = editable hours / rate)","F"); vh=vs+1
    for i,h in enumerate(["Driver","Rate/hr","Cons. hrs","Exp. hrs","Best hrs",""]):
        c=fi.cell(vh,i+1,h); c.font=fnt(9.5,True,WHITE); c.fill=fl(HDR); c.alignment=CTR if i in(1,2,3,4) else LFT; c.border=BD
    dr0=vh+1
    for i,d in enumerate(drivers):
        rr=dr0+i
        fi.cell(rr,1,d["name"]).font=fnt(10); fi.cell(rr,1).alignment=LFT
        cr=fi.cell(rr,2,d["rate"]); cr.font=fnt(10,True,INP); cr.alignment=CTR; cr.number_format=USDc
        for j,h in enumerate(d["hours"]):
            cc=fi.cell(rr,3+j,h); cc.font=fnt(10,True,INP); cc.alignment=CTR; cc.number_format=HRS
        if d.get("basis"): fi.cell(rr,6,d["basis"]).font=fnt(8.5,False,MUT,True)
        for c0 in range(1,6): fi.cell(rr,c0).border=BD
    drN=dr0+len(drivers)-1
    # scenario results
    sr=drN+2; sec(fi,sr,"SCENARIO RESULTS  →  follow every path","F"); sh=sr+1
    fi.cell(sh,1,"Metric").font=fnt(9.5,True,WHITE); fi.cell(sh,1).fill=fl(HDR); fi.cell(sh,1).border=BD
    fi.cell(sh,2,"").fill=fl(HDR); fi.cell(sh,2).border=BD
    for col,lab in zip((3,4,5),("Conservative","Expected","Best")):
        c=fi.cell(sh,col,lab); c.font=fnt(9.5,True,WHITE); c.fill=fl(HDR); c.alignment=CTR; c.border=BD
    def rrow(r,label,fms,fmt,bold=False,bg=None,fc=INK):
        fi.cell(r,1,label).font=fnt(10,bold,NAVY if bold else INK); fi.cell(r,1).alignment=LFT
        fi.cell(r,2,"").border=BD
        for col,fm in zip((3,4,5),fms):
            cc=fi.cell(r,col,fm); cc.number_format=fmt; cc.font=fnt(10,bold,fc); cc.alignment=RGT
        for c0 in range(1,6):
            fi.cell(r,c0).border=BD
            if bg: fi.cell(r,c0).fill=fl(bg)
    # per-driver value rows
    dval0=sh+1
    for i,d in enumerate(drivers):
        rr=dval0+i; hrrow=dr0+i
        rrow(rr,d["name"]+" ($)",[f"=$B${hrrow}*C{hrrow}",f"=$B${hrrow}*D{hrrow}",f"=$B${hrrow}*E{hrrow}"],USD)
    tv=dval0+len(drivers)
    cols="CDE"
    rrow(tv,"Total monthly value",[f"=SUM({c}{dval0}:{c}{tv-1})" for c in cols],USD,bold=True,bg=ALT)
    pc=tv+1; rrow(pc,"Monthly platform cost",[f"=${GT}"]*3,USD,fc=GREY)
    nm=pc+1; rrow(nm,"Net monthly benefit",[f"{c}{tv}-{c}{pc}" and f"={c}{tv}-{c}{pc}" for c in cols],USD,bold=True)
    roi=nm+1
    if roi_mode=="year1":
        roi_lbl="Year-1 ROI (incl. setup + integration)"
    else:
        roi_lbl="Recurring ROI (net ÷ platform fee, excl. one-time)"
    y1v=roi+1; ny1=roi+2; pb=roi+3   # placeholders; defined below in order
    # We need year1 value & net year1 before ROI(year1). Lay rows: ROI, Year1 value, NetYr1, Payback, 3yr.
    rrow(roi,roi_lbl,
         ([f"=({c}{tv}-{c}{pc})/{c}{pc}" for c in cols] if roi_mode!="year1"
          else [f"={c}{ny1}/${Y1C}" for c in cols]),MULT,bold=True,fc=BLUE)
    rrow(y1v,"Year-1 value (after ramp)",[f"={c}{tv}*(12-${RAMP})" for c in cols],USD)
    rrow(ny1,"Net Year-1 benefit (after ramp)",[f"={c}{y1v}-${Y1C}" for c in cols],USD,bold=True,bg=ALT)
    rrow(pb,"Payback (mo, incl. setup + ramp)",
         [f"=${RAMP}+(${OTT}+${GT}*${RAMP})/{c}{nm}" for c in cols],MO,fc=BLUE)
    tyr=pb+1
    rrow(tyr,"3-year net value",[f"={c}{tv}*(36-${RAMP})-(${GT}*36+${OTT})" for c in cols],USD)
    FI="'Financial Impact'!"
    refs=dict(GT=GT,TV=f"D{tv}",NM=f"D{nm}",ROI=f"D{roi}",NY1=f"D{ny1}",PB=f"D{pb}",
              C_TV=f"C{tv}",E_TV=f"E{tv}",C_PC=f"C{pc}",C_ROI=f"C{roi}",E_ROI=f"E{roi}")

    # ============ DASHBOARD ============
    db=wb.create_sheet("Dashboard",0); db.sheet_view.showGridLines=False
    for col,w in zip("ABCDEFGH",[3,21,16,16,16,16,16,3]): db.column_dimensions[col].width=w
    db.merge_cells("A1:H1"); db.merge_cells("A2:H2"); db.merge_cells("A3:H3")
    db["A1"]=f"{client} × Solum Health"; db["A1"].font=fnt(18,True,WHITE)
    db["A1"].alignment=Alignment(horizontal="right",vertical="center",indent=1)
    db["A2"]=f"Business Case Dashboard  ·  {date}  ·  Private & Confidential"
    db["A2"].font=fnt(10,False,"AEBBD0"); db["A2"].alignment=Alignment(horizontal="right",vertical="center",indent=1)
    fillrow(db,1,1,8,NAVY); fillrow(db,2,1,8,NAVY); fillrow(db,3,1,8,BLUE)
    db.row_dimensions[1].height=30; db.row_dimensions[2].height=17; db.row_dimensions[3].height=4
    try:
        img=XLImage(os.path.join(SKILL_DIR,"assets","solum_logo.png")); ratio=img.width/img.height
        img.height=34; img.width=int(34*ratio); db.add_image(img,"B1")
    except Exception: pass
    db["B5"]="THE BOTTOM LINE  —  Expected case"; db["B5"].font=fnt(12,True,NAVY)
    roi_note="net ÷ platform fee · one-time via payback" if roi_mode!="year1" else "net Yr-1 ÷ Yr-1 investment"
    cards=[("Recurring ROI (excl. one-time)" if roi_mode!="year1" else "Year-1 ROI (incl. setup)",f"={FI}{refs['ROI']}",MULT,roi_note),
           ("Net monthly benefit",f"={FI}{refs['NM']}",USD,"value − platform cost"),
           ("Net Year-1 benefit",f"={FI}{refs['NY1']}",USD,"after setup + ramp"),
           ("Payback period",f"={FI}{refs['PB']}",MO,"recover one-time cost"),
           ("Monthly platform cost",f"={FI}{refs['GT']}",USD,"flat, all-in"),
           ("Total monthly value",f"={FI}{refs['TV']}",USD,"sum of value drivers")]
    pos=[(2,7),(4,7),(6,7),(2,11),(4,11),(6,11)]; eb=Side(style="thin",color="DCE6F4")
    for (lbl,fm,fmt,note),(col,row) in zip(cards,pos):
        cl=get_column_letter(col); cr=get_column_letter(col+1)
        for rr in(row,row+1,row+2): db.merge_cells(f"{cl}{rr}:{cr}{rr}")
        db[f"{cl}{row}"]=lbl; db[f"{cl}{row}"].font=fnt(9,True,"5B6B85"); db[f"{cl}{row}"].alignment=Alignment(horizontal="left",indent=1,vertical="center")
        vc=db[f"{cl}{row+1}"]; vc.value=fm; vc.number_format=fmt; vc.font=fnt(22,True,NAVY); vc.alignment=Alignment(horizontal="left",indent=1,vertical="center")
        nc=db[f"{cl}{row+2}"]; nc.value=note; nc.font=fnt(8.5,False,BLUE,True); nc.alignment=Alignment(horizontal="left",indent=1,vertical="top")
        db.row_dimensions[row].height=15; db.row_dimensions[row+1].height=30; db.row_dimensions[row+2].height=14
        for rr in(row,row+1,row+2):
            for cc in(col,col+1): db.cell(rr,cc).fill=fl(WHITE); db.cell(rr,cc).border=Border(eb,eb,eb,eb)
    db["B15"]="SCENARIO PATHS"; db["B15"].font=fnt(11,True,NAVY)
    roi_h="ROI (recur.)" if roi_mode!="year1" else "ROI (Yr-1)"
    for i,h in enumerate(["Path","Monthly value","Monthly cost",roi_h]):
        c=db.cell(16,2+i,h); c.font=fnt(9.5,True,WHITE); c.fill=fl(NAVY); c.alignment=CTR if i else LFT; c.border=BD
    paths=[("Conservative","C"),("Expected","D"),("Best","E")]
    for i,(nm2,cl) in enumerate(paths):
        rr=17+i; a=ALT if i%2 else WHITE
        db.cell(rr,2,nm2).font=fnt(10,True); db.cell(rr,2).alignment=LFT
        db.cell(rr,3,f"={FI}{cl}{tv}").number_format=USD; db.cell(rr,3).alignment=RGT
        db.cell(rr,4,f"={FI}{cl}{pc}").number_format=USD; db.cell(rr,4).alignment=RGT
        db.cell(rr,5,f"={FI}{cl}{roi}").number_format=MULT; db.cell(rr,5).alignment=RGT; db.cell(rr,5).font=fnt(10,True,BLUE)
        for c0 in range(2,6): db.cell(rr,c0).border=BD; db.cell(rr,c0).fill=fl(a)
    ch=BarChart(); ch.type="col"; ch.style=10; ch.title="Monthly Value vs Cost — All Paths"; ch.height=6.2; ch.width=12
    data=Reference(db,min_col=3,max_col=4,min_row=16,max_row=19); cats=Reference(db,min_col=2,min_row=17,max_row=19)
    ch.add_data(data,titles_from_data=True); ch.set_categories(cats); ch.y_axis.numFmt='"$"#,##0'; ch.y_axis.majorGridlines=None
    db.add_chart(ch,"B22")
    if cfg.get("benefits"):
        db["B38"]="WHY IT WORKS"; db["B38"].font=fnt(11,True,NAVY)
        for i,b in enumerate(cfg["benefits"][:3]):
            rr=39+i; db.cell(rr,2,b["stat"]).font=fnt(12,True,BLUE); db.cell(rr,2).alignment=LFT
            db.merge_cells(f"C{rr}:G{rr}"); db.cell(rr,3,f"{b['title']} — {b['desc']}").font=fnt(10,False,INK); db.cell(rr,3).alignment=LFT
    for r0 in range(4,44):
        for c0 in range(1,9):
            cell=db.cell(r0,c0)
            if cell.fill.start_color.rgb in(None,"00000000","FFFFFFFF"): cell.fill=fl(BG)

    # ============ TIMELINE ============
    if cfg.get("timeline"):
        tl=wb.create_sheet("Timeline"); tl.sheet_view.showGridLines=False
        for c,w in zip("ABCDEFGHIJKL",[26,30,12,4.5,4.5,4.5,4.5,4.5,4.5,4.5,4.5,4.5]): tl.column_dimensions[c].width=w
        band(tl,"Suggested Implementation Timeline","From signature to go-live with a validation window built in","L")
        for i,h in enumerate(["Phase","Key activities","Lead"]):
            c=tl.cell(5,i+1,h); c.font=fnt(9.5,True,WHITE); c.fill=fl(HDR); c.alignment=LFT if i<2 else CTR; c.border=BD
        for i in range(9):
            c=tl.cell(5,4+i,f"W{i}"); c.font=fnt(8.5,True,WHITE); c.fill=fl(HDR); c.alignment=CTR; c.border=BD
        for i,ph in enumerate(cfg["timeline"]):
            rr=6+i; a=ALT if i%2 else WHITE
            tl.cell(rr,1,ph["phase"]).font=fnt(9.5,True,NAVY); tl.cell(rr,1).alignment=LFT
            tl.cell(rr,2,ph.get("activities","")).font=fnt(8.5,False,INK); tl.cell(rr,2).alignment=LFT
            tl.cell(rr,3,ph.get("lead","")).font=fnt(8.5,False,MUT); tl.cell(rr,3).alignment=CTR
            for c0 in range(1,4): tl.cell(rr,c0).border=BD; tl.cell(rr,c0).fill=fl(a)
            s,e=ph.get("start",0),ph.get("end",0); col=TEAL if ph.get("kind")=="validation" else BLUE
            for wk in range(9):
                cc=tl.cell(rr,4+wk); cc.border=BD; cc.fill=fl(col) if s<=wk<=e else fl(a)
            tl.row_dimensions[rr].height=30
        lr=6+len(cfg["timeline"])+1
        tl.cell(lr,1,"■ Implementation").font=fnt(9,False,BLUE)
        tl.cell(lr,2,"■ Validation window").font=fnt(9,False,TEAL)
        for r0 in range(4,lr+1):
            for c0 in range(1,13):
                cell=tl.cell(r0,c0)
                if cell.fill.start_color.rgb in(None,"00000000","FFFFFFFF"): cell.fill=fl(BG)

    # ============ ASSUMPTIONS ============
    ag=wb.create_sheet("Assumptions & Guidelines"); ag.sheet_view.showGridLines=False
    for col,w in zip("ABCD",[36,15,12,52]): ag.column_dimensions[col].width=w
    band(ag,"Assumptions, Sources & Guidelines","What's confirmed, what's an estimate, and how to use this model","D")
    for i,h in enumerate(["Item","Value","Status / source",""]):
        if i<3:
            c=ag.cell(6,i+1 if i<2 else 4,h); c.font=fnt(9,True,WHITE); c.fill=fl(HDR); c.alignment=CTR; c.border=BD
    rr=7
    for row in cfg.get("assumptions_sources",[]):
        item,val,src=(row+["",""])[:3]
        ag.cell(rr,1,item).font=fnt(9.5); ag.cell(rr,1).alignment=LFT
        c=ag.cell(rr,2,val); c.font=fnt(9.5,True,INP); c.alignment=CTR
        if "estimate" in str(src).lower() or "confirm" in str(src).lower(): c.fill=fl("FFFF00")
        ag.cell(rr,4,src).font=fnt(8.5,False,MUT); ag.cell(rr,4).alignment=LFT
        for c0 in range(1,5): ag.cell(rr,c0).border=BD
        rr+=1
    rr+=1; sec(ag,rr,"GUIDELINES FOR USE","D"); rr+=1
    guides=cfg.get("guidelines",[
        "Cost figures are firm — confirmed pricing. Value figures use client-provided ranges.",
        "All value drivers scale across paths: Conservative = low-end hours, Best = high-end, Expected = midpoint.",
        ("ROI is the recurring return (net value ÷ platform fee, excluding the one-time); the setup + integration is recovered via payback (incl. a go-live ramp)."
         if roi_mode!="year1" else
         "ROI is Year-1 and includes the one-time setup + integration; value begins after a go-live ramp."),
        "Edit any blue cell on the Financial Impact tab — Dashboard, scenarios and chart update automatically.",
        "Confidential — prepared for "+client+". © 2026 Solum Health Technologies Inc.",
    ])
    for g in guides:
        ag.merge_cells(f"A{rr}:D{rr}"); ag.cell(rr,1,"•  "+g).font=fnt(9.5,False,INK); ag.cell(rr,1).alignment=LFT
        ag.row_dimensions[rr].height=16; rr+=1
    for r0 in range(4,rr+1):
        for c0 in range(1,5):
            cell=ag.cell(r0,c0)
            if cell.fill.start_color.rgb in(None,"00000000","FFFFFFFF"): cell.fill=fl(BG)

    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    wb.calculation.fullCalcOnLoad=True
    out=os.path.join(outdir,f"{client.replace(' ','_')}_BusinessCase.xlsx")
    wb.save(out)
    print("SAVED:",out)
    print("Tabs:",wb.sheetnames)

if __name__=="__main__":
    if len(sys.argv)<2: print("usage: build_model.py <config.json>"); sys.exit(1)
    main(sys.argv[1])

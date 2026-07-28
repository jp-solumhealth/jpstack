#!/usr/bin/env python3
"""dcf.py — build a formula-driven DCF model (.xlsx).

Unlevered-FCF DCF with CAPM/WACC, mid-year convention, perpetuity terminal value,
Bear/Base/Bull scenario sheets, an EV->equity bridge, and a live WACC x terminal-growth
sensitivity grid. Every calculation cell is an Excel FORMULA (inputs are blue) so you can
open it and flex assumptions. A hidden _meta sheet stores the raw inputs as JSON so
validate_model.py can re-derive the math independently.

Usage:
  python3 dcf.py --template > inputs.json     # emit an inputs template to fill in
  python3 dcf.py --inputs inputs.json --out NVDA_dcf_2026-06-03.xlsx
  python3 dcf.py --demo --out demo_dcf.xlsx   # build from a sample

All monetary inputs in millions; shares in millions; rates/margins as decimals (0.21 = 21%).
"""
import argparse, json, sys, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DEMO = {
    "ticker": "DEMO", "price": 100.0, "shares": 1000.0, "net_debt": 500.0,
    "revenue_base": 5000.0, "current_ebit_margin": 0.15, "tax_rate": 0.21,
    "da_pct": 0.05, "capex_pct": 0.06, "nwc_pct": 0.10, "years": 5,
    "terminal_growth": 0.025, "rf": 0.043, "erp": 0.05, "beta": 1.10,
    "cost_debt_pretax": 0.05,
    "scenarios": {
        "bear": {"cagr": 0.04, "target_margin": 0.13},
        "base": {"cagr": 0.08, "target_margin": 0.17},
        "bull": {"cagr": 0.12, "target_margin": 0.20},
    },
}

BLUE = Font(color="1F4E78")           # input cells
BOLD = Font(bold=True)
TITLE = Font(bold=True, size=13)
HEADFILL = PatternFill("solid", fgColor="D9E1F2")
PCT = "0.0%"
MM = "#,##0"
USD = "#,##0.00"
THIN = Border(bottom=Side(style="thin", color="BFBFBF"))


def _in(ws, cell, value, fmt=None):
    ws[cell] = value
    ws[cell].font = BLUE
    if fmt:
        ws[cell].number_format = fmt


def _f(ws, cell, formula, fmt=None, bold=False):
    ws[cell] = formula
    if bold:
        ws[cell].font = BOLD
    if fmt:
        ws[cell].number_format = fmt


def build_assumptions(wb, d):
    ws = wb.create_sheet("Assumptions")
    ws.column_dimensions["A"].width = 26
    for c in "BCD":
        ws.column_dimensions[c].width = 14
    ws["A1"] = f"DCF Model — {d['ticker']}"; ws["A1"].font = TITLE
    ws["A2"] = "Inputs are BLUE. Everything else is a formula — edit blue cells and recalc."
    ws["A2"].font = Font(italic=True, size=9, color="808080")

    rows = [
        ("A4", "Current price", "B4", d["price"], USD),
        ("A5", "Diluted shares (M)", "B5", d["shares"], MM),
        ("A6", "Net debt (M)", "B6", d["net_debt"], MM),
        ("A7", "Base revenue (M)", "B7", d["revenue_base"], MM),
        ("A8", "Current EBIT margin", "B8", d["current_ebit_margin"], PCT),
        ("A9", "Tax rate", "B9", d["tax_rate"], PCT),
        ("A10", "D&A % of revenue", "B10", d["da_pct"], PCT),
        ("A11", "Capex % of revenue", "B11", d["capex_pct"], PCT),
        ("A12", "ΔNWC % of Δrevenue", "B12", d["nwc_pct"], PCT),
        ("A13", "Forecast years (N)", "B13", d["years"], "0"),
        ("A14", "Terminal growth g", "B14", d["terminal_growth"], PCT),
        ("A17", "Risk-free (10Y)", "B17", d["rf"], PCT),
        ("A18", "Equity risk premium", "B18", d["erp"], PCT),
        ("A19", "Beta", "B19", d["beta"], "0.00"),
        ("A20", "Pre-tax cost of debt", "B20", d["cost_debt_pretax"], PCT),
    ]
    for lab_c, lab, val_c, val, fmt in rows:
        ws[lab_c] = lab
        _in(ws, val_c, val, fmt)

    ws["A16"] = "WACC inputs"; ws["A16"].font = BOLD
    ws["A21"] = "Market cap E (=price×shares)"; _f(ws, "B21", "=B4*B5", MM)
    ws["A22"] = "Debt D (=max net debt,0)"; _f(ws, "B22", "=MAX(B6,0)", MM)
    ws["A23"] = "Cost of equity (CAPM)"; _f(ws, "B23", "=B17+B19*B18", PCT)
    ws["A24"] = "After-tax cost of debt"; _f(ws, "B24", "=B20*(1-B9)", PCT)
    ws["A25"] = "WACC"; _f(ws, "B25", "=B21/(B21+B22)*B23+B22/(B21+B22)*B24", PCT, bold=True)

    ws["A27"] = "Scenario drivers"; ws["A27"].font = BOLD
    for col, name in (("B", "bear"), ("C", "base"), ("D", "bull")):
        ws[f"{col}27"] = name.capitalize(); ws[f"{col}27"].font = BOLD
        ws[f"{col}27"].fill = HEADFILL
    ws["A28"] = "Revenue CAGR"
    ws["A29"] = "Target EBIT margin (yr N)"
    sc = d["scenarios"]
    for col, name in (("B", "bear"), ("C", "base"), ("D", "bull")):
        _in(ws, f"{col}28", sc[name]["cagr"], PCT)
        _in(ws, f"{col}29", sc[name]["target_margin"], PCT)
    return ws


def build_dcf_sheet(wb, name, cagr_cell, margin_cell, n):
    """One scenario projection. cagr_cell/margin_cell are abs refs into Assumptions."""
    ws = wb.create_sheet(name)
    ws.column_dimensions["A"].width = 20
    last = get_column_letter(1 + n)  # year columns B..last
    ws["A1"] = f"{name} — UFCF DCF"; ws["A1"].font = TITLE

    labels = {3: "Year", 4: "Period (mid-yr)", 5: "Revenue growth", 6: "Revenue",
              7: "EBIT margin", 8: "EBIT", 9: "NOPAT", 10: "(+) D&A", 11: "(−) Capex",
              12: "(−) ΔNWC", 13: "Unlevered FCF", 14: "Discount factor", 15: "PV of UFCF"}
    for r, lab in labels.items():
        ws[f"A{r}"] = lab
        if r in (6, 13, 15):
            ws[f"A{r}"].font = BOLD

    for i in range(n):
        c = get_column_letter(2 + i)       # current column
        p = get_column_letter(1 + i) if i else None  # previous column (None for yr1)
        yr = i + 1
        ws[f"{c}3"] = yr; ws[f"{c}3"].font = BOLD; ws[f"{c}3"].fill = HEADFILL
        _f(ws, f"{c}4", f"={c}3-0.5", "0.0")
        _f(ws, f"{c}5", f"={cagr_cell}", PCT)
        if i == 0:
            _f(ws, f"{c}6", f"=Assumptions!$B$7*(1+{c}5)", MM)
            _f(ws, f"{c}12", f"=({c}6-Assumptions!$B$7)*Assumptions!$B$12", MM)
        else:
            _f(ws, f"{c}6", f"={p}6*(1+{c}5)", MM)
            _f(ws, f"{c}12", f"=({c}6-{p}6)*Assumptions!$B$12", MM)
        # margin ramps linearly from current (B8) to target (margin_cell) over N years
        _f(ws, f"{c}7", f"=Assumptions!$B$8+({margin_cell}-Assumptions!$B$8)*{c}3/Assumptions!$B$13", PCT)
        _f(ws, f"{c}8", f"={c}6*{c}7", MM)
        _f(ws, f"{c}9", f"={c}8*(1-Assumptions!$B$9)", MM)
        _f(ws, f"{c}10", f"={c}6*Assumptions!$B$10", MM)
        _f(ws, f"{c}11", f"={c}6*Assumptions!$B$11", MM)
        _f(ws, f"{c}13", f"={c}9+{c}10-{c}11-{c}12", MM)
        _f(ws, f"{c}14", f"=1/(1+Assumptions!$B$25)^{c}4", "0.000")
        _f(ws, f"{c}15", f"={c}13*{c}14", MM)

    # Bridge (column A label / B value)
    bridge = {
        17: ("Sum PV (explicit)", f"=SUM(B15:{last}15)", MM),
        18: ("Terminal value", f"={last}13*(1+Assumptions!$B$14)/(Assumptions!$B$25-Assumptions!$B$14)", MM),
        19: ("PV of terminal value", f"=B18/(1+Assumptions!$B$25)^{last}4", MM),
        20: ("Enterprise value", "=B17+B19", MM),
        21: ("(−) Net debt", "=Assumptions!$B$6", MM),
        22: ("Equity value", "=B20-B21", MM),
        23: ("Diluted shares (M)", "=Assumptions!$B$5", MM),
        24: ("Value per share", "=B22/B23", USD),
        25: ("Current price", "=Assumptions!$B$4", USD),
        26: ("Upside / (downside)", "=B24/B25-1", PCT),
        27: ("TV % of EV", "=B19/B20", PCT),
    }
    for r, (lab, formula, fmt) in bridge.items():
        ws[f"A{r}"] = lab
        bold = r in (20, 22, 24, 26)
        _f(ws, f"B{r}", formula, fmt, bold=bold)
        if r in (24, 26):
            ws[f"B{r}"].fill = HEADFILL
    return ws


def build_summary(wb, ticker):
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 18
    for c in "BCD":
        ws.column_dimensions[c].width = 14
    ws["A1"] = f"Valuation Summary — {ticker}"; ws["A1"].font = TITLE
    for c, h in (("A", "Scenario"), ("B", "Value/share"), ("C", "Upside"), ("D", "TV % EV")):
        ws[f"{c}3"] = h; ws[f"{c}3"].font = BOLD; ws[f"{c}3"].fill = HEADFILL
    for i, sc in enumerate(("Bear", "Base", "Bull")):
        r = 4 + i
        ws[f"A{r}"] = sc
        _f(ws, f"B{r}", f"={sc}!B24", USD, bold=(sc == "Base"))
        _f(ws, f"C{r}", f"={sc}!B26", PCT)
        _f(ws, f"D{r}", f"={sc}!B27", PCT)
    ws["A7"] = "Current price"; _f(ws, "B7", "=Assumptions!B4", USD)
    ws["A8"] = "WACC (base)"; _f(ws, "B8", "=Assumptions!B25", PCT)
    ws["A9"] = "Terminal g"; _f(ws, "B9", "=Assumptions!B14", PCT)
    ws["A11"] = "Reconcile DCF range with comps (separate). Weight by forecast confidence."
    ws["A11"].font = Font(italic=True, size=9, color="808080")
    return ws


def build_sensitivity(wb, n):
    """WACC (rows) x terminal-growth (cols) -> base-case value/share, live formulas."""
    ws = wb.create_sheet("Sensitivity")
    last = get_column_letter(1 + n)
    ws["A1"] = "Sensitivity — base value/share: WACC (down) × terminal g (across)"
    ws["A1"].font = TITLE
    ws["A2"] = "Corner cells where g ≥ WACC are not meaningful."
    ws["A2"].font = Font(italic=True, size=9, color="808080")
    # axis values, centered on base, live formulas
    ws["A3"] = "WACC \\ g"; ws["A3"].font = BOLD
    g_off = [-0.01, -0.005, 0.0, 0.005, 0.01]
    w_off = [-0.02, -0.01, 0.0, 0.01, 0.02]
    for j, go in enumerate(g_off):
        col = get_column_letter(2 + j)
        _f(ws, f"{col}3", f"=Assumptions!$B$14+({go})", PCT, bold=True)
        ws[f"{col}3"].fill = HEADFILL
    for i, wo in enumerate(w_off):
        r = 4 + i
        _f(ws, f"A{r}", f"=Assumptions!$B$25+({wo})", PCT, bold=True)
        ws[f"A{r}"].fill = HEADFILL
    # interior: value/share at (W=$A{r}, g={col}$3) using Base UFCF + period rows
    uf = f"Base!$B$13:${last}$13"
    pr = f"Base!$B$4:${last}$4"
    lastuf = f"Base!${last}$13"
    lastpr = f"Base!${last}$4"
    for i in range(5):
        r = 4 + i
        for j in range(5):
            col = get_column_letter(2 + j)
            W = f"$A{r}"
            g = f"{col}$3"
            formula = (f"=(SUMPRODUCT({uf},1/(1+{W})^{pr})"
                       f"+{lastuf}*(1+{g})/({W}-{g})/(1+{W})^{lastpr}"
                       f"-Assumptions!$B$6)/Assumptions!$B$5")
            _f(ws, f"{col}{r}", formula, USD)
    return ws


def build_meta(wb, d):
    ws = wb.create_sheet("_meta")
    ws["A1"] = json.dumps(d)
    ws.sheet_state = "hidden"


def build(d, out):
    wb = Workbook()
    wb.remove(wb.active)
    n = int(d["years"])
    build_assumptions(wb, d)
    build_dcf_sheet(wb, "Bear", "Assumptions!$B$28", "Assumptions!$B$29", n)
    build_dcf_sheet(wb, "Base", "Assumptions!$C$28", "Assumptions!$C$29", n)
    build_dcf_sheet(wb, "Bull", "Assumptions!$D$28", "Assumptions!$D$29", n)
    build_summary(wb, d["ticker"])
    build_sensitivity(wb, n)
    build_meta(wb, d)
    wb.save(out)


def main():
    ap = argparse.ArgumentParser(description="Build a formula-driven DCF .xlsx.")
    ap.add_argument("--inputs", help="JSON inputs file")
    ap.add_argument("--out", help="output .xlsx path")
    ap.add_argument("--demo", action="store_true", help="use built-in sample inputs")
    ap.add_argument("--template", action="store_true", help="print an inputs JSON template and exit")
    args = ap.parse_args()

    if args.template:
        print(json.dumps(DEMO, indent=2))
        return
    if args.demo:
        d = DEMO
    elif args.inputs:
        with open(args.inputs) as f:
            d = json.load(f)
    else:
        ap.error("provide --inputs FILE, or --demo, or --template")

    today = datetime.date.today().isoformat()
    out = args.out or f"{d.get('ticker','model')}_dcf_{today}.xlsx"
    build(d, out)
    print(f"Built {out}  ({d['ticker']}, {d['years']}y, Bear/Base/Bull + sensitivity).")
    print("Next: run validate_model.py on it, then reconcile with comps before any rating.")


if __name__ == "__main__":
    main()

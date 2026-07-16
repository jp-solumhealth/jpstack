#!/usr/bin/env python3
"""Buy-vs-Build underwriting workbook generator.

Builds a 13-tab, fully formula-linked Excel model comparing Build-to-Sell vs
Build-to-Hold for small residential development (duplex/quadplex/SFR).
All assumptions live on the Assumptions tab; `Assumptions!Units` (2 or 4)
re-scales cost, fees, rent, revenue and financing in one cell.

Usage:
    python3 build_model.py --out model.xlsx
    python3 build_model.py --config deal.json --out model.xlsx

deal.json overrides any DEFAULT_CONFIG key (defaults = worked Marion County FL
duplex example). After building, ALWAYS run validate_model.py on the output.
"""
import argparse, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

DEFAULT_CONFIG = {
    "title": "MULTIFAMILY DEVELOPMENT MODEL",
    "subtitle": "Buy-vs-Build underwriting",
    # product
    "units": 2, "total_sqft": 2010, "ac_sqft": 1904,
    "location": "Target submarket",
    # land
    "land": 60000, "land_close_pct": 0.02,
    # contractor budget (2-unit basis; appliances are PER UNIT)
    "budget_sub": 200045, "appl_unit": 2500, "toilet": 750,
    "oh_pct": 0.02, "fee_pct": 0.09,
    # soft costs & county fees
    "plans_design": 8000, "bldg_permit": 2500, "impact_door": 9400,
    "utility_conn": 0, "contingency_pct": 0.05,
    "tax_constr": 1200, "ins_constr": 2800, "legal_misc": 2000,
    # construction financing
    "c_rate": 0.0875, "c_ltc": 0.75, "c_term": 12,
    "c_orig_pts": 0.02, "c_avg_draw": 0.60,
    # build-to-sell
    "sale_unit": 192500, "comm_pct": 0.05, "sale_close_pct": 0.015,
    # build-to-hold
    "rent_unit": 1250, "vacancy_pct": 0.06, "other_inc": 0,
    "mgmt_pct": 0.10, "tax_pct_val": 0.011, "ins_unit": 1750,
    "rm_pct": 0.05, "reserve_unit": 350, "util_admin": 900, "lease_pct": 0.01,
    # exit / refi
    "exit_cap": 0.070, "refi_ltv": 0.75, "refi_rate": 0.0625, "amort_yrs": 30,
    # growth & return
    "rent_growth": 0.03, "exp_infl": 0.03, "apprec": 0.03,
    "term_cap": 0.075, "disc_rate": 0.10, "hold_years": 30,
    # sensitivity grids
    "sens_cost_var": [-0.10, -0.05, 0.0, 0.05, 0.10],
    "sens_sale_var": [-0.10, -0.05, 0.0, 0.05, 0.10],
    "sens_caps": [0.06, 0.065, 0.07, 0.075, 0.08],
    "sens_rents": [1150, 1250, 1400, 1500, 1650],
    # market summary rows: [section, [[metric, value, source], ...]]
    "market_rows": [
        ["FILL FROM MARKET RESEARCH", [
            ["Population / growth", "—", "Census"],
            ["Median sale price / supply / DOM", "—", "MLS"],
            ["Rents by bedroom / growth", "—", "Zumper, RentCafe"],
            ["Cap rates (small MF)", "—", "broker research"],
            ["Zoning / min-lot / impact fees", "—", "county LDC"],
        ]],
    ],
    "recommendation": ("FILL AFTER MODELING: verdict (buy / build-to-sell / "
                       "build-to-hold / wait), the 2-3 needle-movers, and the "
                       "county calls that close open items."),
}

FONT = "Arial"
NAVY = "1F3A5F"; TEAL = "143642"; CLAY = "C57B45"
LGREY = "EEF1F4"; PAPER = "FAF8F4"
BLUE = "0000FF"; BLACK = "000000"; GREEN = "008000"; WHITE = "FFFFFF"
YEL = "FFF3C4"; WARN = "C99A3B"
CUR = '$#,##0;($#,##0);"-"'; CUR2 = '$#,##0.00;($#,##0.00);"-"'
PCT = '0.0%'; MULT = '0.00"x"'; NUM = '#,##0;(#,##0);"-"'
thin = Side(style="thin", color="B8BFC7")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)


def build(C, out_path):
    def cell(ws, coord, val=None, *, bold=False, italic=False, size=10, color=BLACK,
             fill=None, fmt=None, align=None, wrap=False, border=False, valign=None):
        c = ws[coord]
        if val is not None:
            c.value = val
        c.font = Font(name=FONT, bold=bold, italic=italic, size=size, color=color)
        if fill: c.fill = PatternFill("solid", fgColor=fill)
        if fmt: c.number_format = fmt
        if align or wrap or valign:
            c.alignment = Alignment(horizontal=align, wrap_text=wrap, vertical=valign or "center")
        if border: c.border = border_all
        return c

    def title_band(ws, row, text, sub=None, span=8):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        cell(ws, f"A{row}", text, bold=True, size=15, color=WHITE, fill=TEAL, align="left")
        ws.row_dimensions[row].height = 26
        if sub:
            r2 = row + 1
            ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=span)
            cell(ws, f"A{r2}", sub, italic=True, size=9, color="5A6470", fill=PAPER, align="left")
            ws.row_dimensions[r2].height = 16
            return row + 2
        return row + 1

    def section(ws, row, text, span=8, fill=NAVY):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        cell(ws, f"A{row}", text, bold=True, size=11, color=WHITE, fill=fill, align="left")
        ws.row_dimensions[row].height = 20

    wb = Workbook()

    # ---------------- ASSUMPTIONS ----------------
    A = {}
    aw = wb.active; aw.title = "Assumptions"
    aw.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", [40, 16, 10, 40, 4, 4]):
        aw.column_dimensions[col].width = w
    r = title_band(aw, 1, C["title"],
                   "Assumptions & Inputs  |  Blue = input · Black = formula · Green = cross-sheet link · Yellow = key driver", span=4)
    r += 1
    state = {"r": r}

    def arow(key, label, val, fmt=None, style="input", note=""):
        rr = state["r"]
        cell(aw, f"A{rr}", label, size=10)
        col = {"input": BLUE, "formula": BLACK, "link": GREEN, "key": BLUE}[style]
        fill = YEL if style == "key" else None
        cell(aw, f"B{rr}", val, color=col, fmt=fmt, align="right", fill=fill)
        if note: cell(aw, f"D{rr}", note, italic=True, size=8, color="6A7480")
        if key: A[key] = f"'Assumptions'!$B${rr}"
        state["r"] = rr + 1

    def ahead(text):
        section(aw, state["r"], text, span=4, fill=NAVY)
        state["r"] += 1

    ahead("PROJECT / PRODUCT")
    arow("units", "Units (2 = Duplex · 4 = Quadplex)", C["units"], NUM, "key",
         "Drives cost, fees, rent, revenue. Land held constant.")
    arow("total_sqft", "Total building SF", C["total_sqft"], NUM, "input", "Source: contractor budget")
    arow("ac_sqft", "SF under A/C", C["ac_sqft"], NUM, "input")
    arow("unit_sqft", "SF per unit", f"={A['total_sqft']}/{A['units']}", NUM, "formula")
    arow("location", "Target submarket", C["location"], None, "input",
         "Verify zoning + min-lot feasibility before underwriting")

    ahead("LAND & ACQUISITION")
    arow("land", "Land / lot cost", C["land"], CUR, "key")
    arow("land_close_pct", "Land closing costs %", C["land_close_pct"], PCT, "input")
    arow("land_close", "Land closing costs $", f"={A['land']}*{A['land_close_pct']}", CUR, "formula")

    ahead("CONSTRUCTION — HARD COST (contractor budget)")
    arow("budget_sub", "Contractor direct subtotal (2-unit basis)", C["budget_sub"], CUR, "input",
         "Source: contractor budget direct-cost subtotal")
    arow("appl_unit", "Appliances per unit", C["appl_unit"], CUR, "input", "x2 duplex / x4 quad = per-unit")
    arow("toilet", "Portable toilet (per project)", C["toilet"], CUR, "input")
    arow("oh_pct", "Overhead %", C["oh_pct"], PCT, "input")
    arow("fee_pct", "Contractor fee %", C["fee_pct"], PCT, "input")
    arow("hard_perunit", "Hard cost per unit (ex-appliances)",
         f"=({A['budget_sub']}+{A['toilet']}+{A['budget_sub']}*{A['oh_pct']}+{A['budget_sub']}*{A['fee_pct']})/2",
         CUR, "formula", "2-unit budget non-appliance hard / 2; scales linearly")
    arow("appl_total", "Appliances total", f"={A['appl_unit']}*{A['units']}", CUR, "formula")
    arow("hard_total", "TOTAL HARD CONSTRUCTION COST",
         f"={A['hard_perunit']}*{A['units']}+{A['appl_total']}", CUR, "formula",
         "Excludes contractor plans/permits line (reclassified to soft)")

    ahead("SOFT COSTS & COUNTY FEES")
    arow("plans_design", "Plans, architecture, engineering, survey", C["plans_design"], CUR, "input")
    arow("bldg_permit", "Building permit", C["bldg_permit"], CUR, "input")
    arow("impact_door", "Impact fees per door", C["impact_door"], CUR, "key",
         "Confirm county schedule; MF may differ from SFR")
    arow("impact_total", "Impact fees total", f"={A['impact_door']}*{A['units']}", CUR, "formula")
    arow("utility_conn", "Central water/sewer connection", C["utility_conn"], CUR, "input",
         "$0 if budget builds well+septic (avoid double count)")
    arow("contingency_pct", "Contingency % of hard cost", C["contingency_pct"], PCT, "input")
    arow("contingency", "Contingency $", f"={A['hard_total']}*{A['contingency_pct']}", CUR, "formula")
    arow("tax_constr", "Property tax during construction", C["tax_constr"], CUR, "input")
    arow("ins_constr", "Builder's risk insurance", C["ins_constr"], CUR, "input")
    arow("legal_misc", "Legal, closing & misc soft", C["legal_misc"], CUR, "input")

    ahead("DEVELOPMENT COST ROLL-UP")
    arow("tdc_exfin", "Total Development Cost (ex-financing)",
         f"={A['land']}+{A['land_close']}+{A['hard_total']}+{A['plans_design']}+{A['bldg_permit']}"
         f"+{A['impact_total']}+{A['utility_conn']}+{A['contingency']}+{A['tax_constr']}+{A['ins_constr']}+{A['legal_misc']}",
         CUR, "formula")

    ahead("CONSTRUCTION FINANCING")
    arow("c_rate", "Construction loan rate", C["c_rate"], PCT, "key")
    arow("c_ltc", "Loan-to-Cost", C["c_ltc"], PCT, "input")
    arow("c_term", "Construction + lease-up term (months)", C["c_term"], NUM, "input")
    arow("c_orig_pts", "Origination points", C["c_orig_pts"], PCT, "input")
    arow("c_avg_draw", "Avg outstanding balance %", C["c_avg_draw"], PCT, "input", "S-curve draw")
    arow("c_loan", "Construction loan amount", f"={A['tdc_exfin']}*{A['c_ltc']}", CUR, "formula")
    arow("c_orig_fee", "Origination fee", f"={A['c_loan']}*{A['c_orig_pts']}", CUR, "formula")
    arow("c_interest", "Construction interest (accrued)",
         f"={A['c_loan']}*{A['c_rate']}*({A['c_term']}/12)*{A['c_avg_draw']}", CUR, "formula")
    arow("c_fin_cost", "Total construction financing cost", f"={A['c_orig_fee']}+{A['c_interest']}", CUR, "formula")
    arow("tpc", "TOTAL PROJECT COST (all-in)", f"={A['tdc_exfin']}+{A['c_fin_cost']}", CUR, "formula")
    arow("equity_constr", "Equity required (construction)", f"={A['tpc']}-{A['c_loan']}", CUR, "formula",
         "TPC includes financing — do NOT add financing to equity again")
    arow("cost_unit", "Cost per unit", f"={A['tpc']}/{A['units']}", CUR, "formula")
    arow("cost_sf", "Cost per building SF", f"={A['tpc']}/{A['total_sqft']}", CUR2, "formula")

    ahead("BUILD-TO-SELL — REVENUE")
    arow("sale_unit", "Sale price per unit", C["sale_unit"], CUR, "key", "From retail comps, not distressed")
    arow("gross_sales", "Gross sales revenue", f"={A['sale_unit']}*{A['units']}", CUR, "formula")
    arow("sale_psf", "Implied sale $/SF", f"={A['gross_sales']}/{A['total_sqft']}", CUR2, "formula")
    arow("comm_pct", "Sales commission %", C["comm_pct"], PCT, "input")
    arow("sale_close_pct", "Seller closing costs %", C["sale_close_pct"], PCT, "input")
    arow("sale_costs", "Total selling costs", f"={A['gross_sales']}*({A['comm_pct']}+{A['sale_close_pct']})", CUR, "formula")
    arow("net_sales", "Net sales revenue", f"={A['gross_sales']}-{A['sale_costs']}", CUR, "formula")

    ahead("BUILD-TO-HOLD — RENTAL")
    arow("rent_unit", "Monthly rent per unit", C["rent_unit"], CUR, "key", "Sensitize: biggest lever")
    arow("gsr", "Gross scheduled rent (annual)", f"={A['rent_unit']}*{A['units']}*12", CUR, "formula")
    arow("vacancy_pct", "Vacancy & credit loss %", C["vacancy_pct"], PCT, "input")
    arow("other_inc", "Other income (annual)", C["other_inc"], CUR, "input")
    arow("mgmt_pct", "Property management % of EGI", C["mgmt_pct"], PCT, "key")
    arow("tax_pct_val", "Property tax % (on cost basis)", C["tax_pct_val"], PCT, "input")
    arow("ins_unit", "Insurance per unit / yr", C["ins_unit"], CUR, "input")
    arow("rm_pct", "Repairs & maintenance % of EGI", C["rm_pct"], PCT, "input")
    arow("reserve_unit", "Capital reserve per unit", C["reserve_unit"], CUR, "input")
    arow("util_admin", "Owner utilities + admin (annual)", C["util_admin"], CUR, "input")
    arow("lease_pct", "Leasing / advertising % of EGI", C["lease_pct"], PCT, "input")

    ahead("EXIT / REFINANCE")
    arow("exit_cap", "Stabilized cap rate", C["exit_cap"], PCT, "key")
    arow("refi_ltv", "DSCR refinance LTV", C["refi_ltv"], PCT, "key")
    arow("refi_rate", "DSCR loan rate", C["refi_rate"], PCT, "key")
    arow("amort_yrs", "Amortization (years)", C["amort_yrs"], NUM, "input")

    ahead("GROWTH & RETURN")
    arow("rent_growth", "Annual rent growth", C["rent_growth"], PCT, "input")
    arow("exp_infl", "Operating expense inflation", C["exp_infl"], PCT, "input")
    arow("apprec", "Annual appreciation", C["apprec"], PCT, "input")
    arow("term_cap", "Terminal cap (final yr)", C["term_cap"], PCT, "input")
    arow("disc_rate", "Discount rate (NPV)", C["disc_rate"], PCT, "input")
    arow("hold_years", "Hold period (years)", C["hold_years"], NUM, "input")
    aw.freeze_panes = "A3"

    # ---------------- DEVELOPMENT BUDGET ----------------
    db = wb.create_sheet("Development Budget")
    db.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", [46, 16, 16, 34, 2]):
        db.column_dimensions[col].width = w
    r = title_band(db, 1, "DEVELOPMENT BUDGET", "Contractor basis reconciled to institutional all-in", span=4)
    r += 1
    cell(db, f"A{r}", "Line item", bold=True, color=WHITE, fill=NAVY, border=True)
    cell(db, f"B{r}", "Amount", bold=True, color=WHITE, fill=NAVY, align="right", border=True)
    cell(db, f"C{r}", "Basis", bold=True, color=WHITE, fill=NAVY, align="right", border=True)
    cell(db, f"D{r}", "Source / note", bold=True, color=WHITE, fill=NAVY, border=True)
    r += 1
    st = {"r": r}

    def dbrow(label, val, basis="", note="", bold=False, fill=None, fmt=CUR):
        rr = st["r"]
        cell(db, f"A{rr}", label, bold=bold, border=True, fill=fill)
        cell(db, f"B{rr}", val, color=GREEN, fmt=fmt, align="right", border=True, bold=bold, fill=fill)
        cell(db, f"C{rr}", basis or None, align="right", border=True, fill=fill, size=9, color="6A7480")
        cell(db, f"D{rr}", note, italic=True, size=8, color="6A7480", border=True, fill=fill)
        st["r"] = rr + 1

    section(db, st["r"], "A. CONTRACTOR BUDGET (as submitted)", span=4, fill=TEAL); st["r"] += 1
    dbrow("Direct construction subtotal", f"={A['budget_sub']}")
    dbrow("Appliances (per-unit × Units)", f"={A['appl_total']}", "×Units")
    dbrow("Portable toilet", f"={A['toilet']}")
    dbrow("Overhead", f"={A['budget_sub']}*{A['oh_pct']}")
    dbrow("Contractor fee", f"={A['budget_sub']}*{A['fee_pct']}")
    dbrow("Contractor total (corrected)",
          f"={A['budget_sub']}+{A['appl_total']}+{A['toilet']}+{A['budget_sub']}*{A['oh_pct']}+{A['budget_sub']}*{A['fee_pct']}",
          "", "Check the sheet's own TOTAL for omitted OH/fee rows", bold=True, fill=LGREY)
    section(db, st["r"], "B. INSTITUTIONAL ALL-IN", span=4, fill=TEAL); st["r"] += 1
    dbrow("Land / lot", f"={A['land']}")
    dbrow("Land closing costs", f"={A['land_close']}")
    dbrow("Hard construction cost", f"={A['hard_total']}")
    dbrow("Plans, engineering, survey", f"={A['plans_design']}")
    dbrow("Building permit", f"={A['bldg_permit']}")
    dbrow("Impact fees (county)", f"={A['impact_total']}", "×Units")
    dbrow("Utility connection", f"={A['utility_conn']}")
    dbrow("Contingency", f"={A['contingency']}")
    dbrow("Property tax during construction", f"={A['tax_constr']}")
    dbrow("Builder's risk insurance", f"={A['ins_constr']}")
    dbrow("Legal & misc", f"={A['legal_misc']}")
    dbrow("Construction interest", f"={A['c_interest']}")
    dbrow("Loan origination", f"={A['c_orig_fee']}")
    dbrow("TOTAL PROJECT COST (all-in)", f"={A['tpc']}", "", "", bold=True, fill="D6E4D0")
    dbrow("Cost per unit", f"={A['cost_unit']}", bold=True)
    dbrow("Cost per building SF", f"={A['cost_sf']}", bold=True, fmt=CUR2)
    db.freeze_panes = "A4"

    # ---------------- BUILD-TO-HOLD ----------------
    bh = wb.create_sheet("Build-to-Hold")
    bh.sheet_view.showGridLines = False
    for col, w in zip("ABCD", [40, 18, 40, 2]):
        bh.column_dimensions[col].width = w
    r = title_band(bh, 1, "BUILD-TO-HOLD — STABILIZED YEAR 1", "Refinance into DSCR loan at stabilization", span=3)
    r += 1
    st = {"r": r}

    def bhrow(label, val, note="", bold=False, fmt=CUR, fill=None, pct=False, mult=False):
        rr = st["r"]
        cell(bh, f"A{rr}", label, bold=bold, fill=fill)
        f = PCT if pct else (MULT if mult else fmt)
        cell(bh, f"B{rr}", val, color=GREEN, fmt=f, align="right", bold=bold, fill=fill)
        if note: cell(bh, f"C{rr}", note, italic=True, size=8, color="6A7480")
        st["r"] = rr + 1
        return rr

    section(bh, st["r"], "INCOME", span=3, fill=NAVY); st["r"] += 1
    r_gsr = bhrow("Gross Scheduled Rent", f"={A['gsr']}")
    r_vac = bhrow("Vacancy & credit loss", f"=-{A['gsr']}*{A['vacancy_pct']}")
    r_oi = bhrow("Other income", f"={A['other_inc']}")
    r_egi = bhrow("Effective Gross Income (EGI)", f"=B{r_gsr}+B{r_vac}+B{r_oi}", bold=True)
    section(bh, st["r"], "OPERATING EXPENSES", span=3, fill=NAVY); st["r"] += 1
    r_o1 = bhrow("Property management", f"=-B{r_egi}*{A['mgmt_pct']}")
    bhrow("Property taxes", f"=-{A['tax_pct_val']}*{A['tpc']}", "on cost basis (non-circular)")
    bhrow("Insurance", f"=-{A['ins_unit']}*{A['units']}")
    bhrow("Repairs & maintenance", f"=-B{r_egi}*{A['rm_pct']}")
    bhrow("Capital reserve", f"=-{A['reserve_unit']}*{A['units']}")
    bhrow("Utilities / admin", f"=-{A['util_admin']}")
    r_o7 = bhrow("Leasing / advertising", f"=-B{r_egi}*{A['lease_pct']}")
    r_opx = bhrow("Total Operating Expenses", f"=SUM(B{r_o1}:B{r_o7})", bold=True)
    r_noi = bhrow("NET OPERATING INCOME (NOI)", f"=B{r_egi}+B{r_opx}", bold=True, fill="D6E4D0")
    bhrow("Operating expense ratio", f"=-B{r_opx}/B{r_egi}", pct=True)
    r_val = bhrow("Stabilized value (NOI / cap)", f"=B{r_noi}/{A['exit_cap']}", bold=True, fill=YEL)
    section(bh, st["r"], "DSCR REFINANCE", span=3, fill=NAVY); st["r"] += 1
    bhrow("Refinance LTV", f"={A['refi_ltv']}", pct=True)
    r_refi = bhrow("Refinance loan proceeds", f"=B{r_val}*{A['refi_ltv']}", bold=True)
    r_pay = bhrow("Construction loan payoff", f"=-{A['c_loan']}")
    r_cash = bhrow("Cash-out (+) / cash-in (−) at refi", f"=B{r_refi}+B{r_pay}", "negative = must inject cash")
    bhrow("DSCR rate", f"={A['refi_rate']}", pct=True)
    r_pmt = bhrow("Monthly debt service", f"=PMT({A['refi_rate']}/12,{A['amort_yrs']}*12,-B{r_refi})", bold=True)
    r_ads = bhrow("Annual debt service", f"=B{r_pmt}*12", bold=True)
    r_dscr = bhrow("DSCR", f"=B{r_noi}/B{r_ads}", bold=True, fill=YEL, mult=True)
    r_cfbt = bhrow("Cash flow before tax (Yr1)", f"=B{r_noi}-B{r_ads}", bold=True)
    r_capr = bhrow("Yield on cost (NOI/TPC)", f"=B{r_noi}/{A['tpc']}", "must beat market cap or hold destroys value", pct=True)
    r_eqrem = bhrow("Equity remaining after refi", f"={A['equity_constr']}-B{r_cash}")
    r_coc = bhrow("Cash-on-cash (Yr1)", f"=IF(B{r_eqrem}<=0,\"n/a (all equity out)\",B{r_cfbt}/B{r_eqrem})", bold=True, fill=YEL, pct=True)
    bh.freeze_panes = "A4"
    BH = {k: f"'Build-to-Hold'!$B${v}" for k, v in
          dict(noi=r_noi, val=r_val, refiloan=r_refi, ads=r_ads, cfbt=r_cfbt,
               eqrem=r_eqrem, cashout=r_cash, egi=r_egi, opxtot=r_opx, pmt=r_pmt,
               dscr=r_dscr, coc=r_coc, capr=r_capr).items()}

    # ---------------- CONSTRUCTION LOAN ----------------
    cl = wb.create_sheet("Construction Loan Model")
    cl.sheet_view.showGridLines = False
    for col, w in zip("ABCD", [40, 18, 40, 2]):
        cl.column_dimensions[col].width = w
    r = title_band(cl, 1, "CONSTRUCTION LOAN MODEL", "Interest-only, S-curve draws, interest reserve", span=3)
    r += 1
    st = {"r": r}

    def clrow(label, val, note="", bold=False, fmt=CUR, fill=None, pct=False):
        rr = st["r"]
        cell(cl, f"A{rr}", label, bold=bold, fill=fill)
        cell(cl, f"B{rr}", val, color=GREEN, fmt=(PCT if pct else fmt), align="right", bold=bold, fill=fill)
        if note: cell(cl, f"C{rr}", note, italic=True, size=8, color="6A7480")
        st["r"] = rr + 1

    section(cl, st["r"], "LOAN SIZING", span=3, fill=NAVY); st["r"] += 1
    clrow("Total Project Cost (all-in)", f"={A['tpc']}", bold=True)
    clrow("Loan-to-Cost", f"={A['c_ltc']}", pct=True)
    clrow("Construction loan amount", f"={A['c_loan']}", bold=True)
    clrow("Equity required", f"={A['equity_constr']}", bold=True, fill=YEL)
    clrow("Stabilized value", f"={BH['val']}")
    clrow("Loan-to-VALUE (red flag if >100%)", f"={A['c_loan']}/{BH['val']}", "lenders size to ARV, not cost", pct=True)
    section(cl, st["r"], "INTEREST & FEES", span=3, fill=NAVY); st["r"] += 1
    clrow("Interest rate", f"={A['c_rate']}", pct=True)
    clrow("Term (months)", f"={A['c_term']}", fmt=NUM)
    clrow("Avg outstanding balance %", f"={A['c_avg_draw']}", pct=True)
    clrow("Accrued interest (reserve)", f"={A['c_interest']}", bold=True)
    clrow("Origination fee", f"={A['c_orig_fee']}")
    clrow("Total financing cost", f"={A['c_fin_cost']}", bold=True)
    cl.freeze_panes = "A4"

    # ---------------- BUILD-TO-SELL ----------------
    bs = wb.create_sheet("Build-to-Sell")
    bs.sheet_view.showGridLines = False
    for col, w in zip("ABCD", [40, 18, 40, 2]):
        bs.column_dimensions[col].width = w
    r = title_band(bs, 1, "BUILD-TO-SELL ANALYSIS", "Develop and sell at completion", span=3)
    r += 1
    st = {"r": r}

    def bsrow(label, val, note="", bold=False, fmt=CUR, fill=None, pct=False, mult=False):
        rr = st["r"]
        cell(bs, f"A{rr}", label, bold=bold, fill=fill)
        f = PCT if pct else (MULT if mult else fmt)
        cell(bs, f"B{rr}", val, color=GREEN, fmt=f, align="right", bold=bold, fill=fill)
        if note: cell(bs, f"C{rr}", note, italic=True, size=8, color="6A7480")
        st["r"] = rr + 1
        return rr

    section(bs, st["r"], "COST & EQUITY", span=3, fill=NAVY); st["r"] += 1
    bsrow("Total Project Cost (all-in)", f"={A['tpc']}", bold=True)
    bsrow("Cost per unit", f"={A['cost_unit']}")
    bsrow("Cost per SF", f"={A['cost_sf']}", fmt=CUR2)
    r_eq = bsrow("Equity invested", f"={A['equity_constr']}", "TPC − loan (financing already in TPC)", bold=True, fill=YEL)
    section(bs, st["r"], "REVENUE", span=3, fill=NAVY); st["r"] += 1
    bsrow("Gross sales revenue", f"={A['gross_sales']}", bold=True)
    bsrow("Implied $/SF", f"={A['sale_psf']}", fmt=CUR2)
    bsrow("Less: selling costs", f"=-{A['sale_costs']}")
    bsrow("Net sales revenue", f"={A['net_sales']}", bold=True)
    section(bs, st["r"], "PROFIT & RETURNS", span=3, fill=NAVY); st["r"] += 1
    r_gp = bsrow("Gross profit", f"={A['gross_sales']}-{A['tpc']}", bold=True)
    r_np = bsrow("Net profit", f"={A['net_sales']}-{A['tpc']}", bold=True, fill="D6E4D0")
    bsrow("Profit margin (on revenue)", f"=B{r_np}/{A['gross_sales']}", pct=True)
    r_roc = bsrow("Return on cost", f"=B{r_np}/{A['tpc']}", bold=True, fill=YEL, pct=True)
    bsrow("Equity multiple", f"=(B{r_eq}+B{r_np})/B{r_eq}", mult=True, bold=True)
    bsrow("Profit per unit", f"=B{r_np}/{A['units']}")
    r_be = bsrow("Breakeven sale price (total)", f"={A['tpc']}/(1-{A['comm_pct']}-{A['sale_close_pct']})", bold=True)
    bsrow("Breakeven per unit", f"=B{r_be}/{A['units']}")
    section(bs, st["r"], "IRR (monthly cash flows over term)", span=3, fill=NAVY); st["r"] += 1
    rr = st["r"]
    cell(bs, f"A{rr}", "Month", bold=True, fill=LGREY)
    for m in range(0, 13):
        cell(bs, get_column_letter(2 + m) + str(rr), m, bold=True, fill=LGREY, align="right", fmt=NUM)
    st["r"] += 1
    r_cf = st["r"]
    cell(bs, f"A{r_cf}", "Equity cash flow", bold=True)
    cell(bs, f"B{r_cf}", f"=-B{r_eq}", fmt=CUR, align="right")
    for m in range(1, 12):
        cell(bs, get_column_letter(2 + m) + str(r_cf), 0, fmt=CUR, align="right")
    cell(bs, get_column_letter(14) + str(r_cf), f"=B{r_eq}+B{r_np}", fmt=CUR, align="right")
    st["r"] += 1
    r_mirr = bsrow("Monthly IRR", f"=IRR(B{r_cf}:N{r_cf})", pct=True)
    r_airr = bsrow("Annualized IRR (project)", f"=(1+B{r_mirr})^12-1", bold=True, fill=YEL, pct=True)
    bs.freeze_panes = "A4"
    BS = {"np": f"'Build-to-Sell'!$B${r_np}", "eq": f"'Build-to-Sell'!$B${r_eq}",
          "roc": f"'Build-to-Sell'!$B${r_roc}", "irr": f"'Build-to-Sell'!$B${r_airr}"}

    # ---------------- P&L YEAR 1 ----------------
    pl = wb.create_sheet("P&L Year 1")
    pl.sheet_view.showGridLines = False
    for col, w in zip("ABC", [42, 18, 40]):
        pl.column_dimensions[col].width = w
    r = title_band(pl, 1, "PRO-FORMA OPERATING STATEMENT — YEAR 1", "Stabilized rental operations", span=3)
    r += 1
    st = {"r": r}

    def plrow(label, ref, bold=False, fill=None, pct=False, fmt=CUR):
        rr = st["r"]
        cell(pl, f"A{rr}", label, bold=bold, fill=fill)
        cell(pl, f"B{rr}", ref, color=GREEN, fmt=(PCT if pct else fmt), align="right", bold=bold, fill=fill)
        st["r"] = rr + 1

    plrow("Gross Scheduled Rent", f"={A['gsr']}")
    plrow("Vacancy & Credit Loss", f"=-{A['gsr']}*{A['vacancy_pct']}")
    plrow("Effective Gross Income", f"={BH['egi']}", bold=True)
    plrow("Total Operating Expenses", f"={BH['opxtot']}")
    plrow("NET OPERATING INCOME", f"={BH['noi']}", bold=True, fill="D6E4D0")
    plrow("Annual Debt Service", f"=-{BH['ads']}")
    plrow("Cash Flow Before Tax", f"={BH['cfbt']}", bold=True, fill=YEL)
    plrow("DSCR", f"={BH['dscr']}", bold=True, fmt=MULT)
    plrow("Yield on cost", f"={BH['capr']}", pct=True)
    plrow("Cash-on-Cash", f"={BH['coc']}", bold=True, pct=True)
    pl.freeze_panes = "A4"

    # ---------------- 30-YEAR PROJECTION ----------------
    pr = wb.create_sheet("30-Year Projection")
    pr.sheet_view.showGridLines = False
    pr.column_dimensions["A"].width = 30
    for i in range(31):
        pr.column_dimensions[get_column_letter(2 + i)].width = 12
    r = title_band(pr, 1, "LONG-HOLD PROJECTION", "Refi at stabilization; sale at terminal cap in final year", span=12)
    r += 1
    hdr = r
    cell(pr, f"A{r}", "Year", bold=True, color=WHITE, fill=NAVY, border=True)
    for y in range(0, 31):
        cell(pr, get_column_letter(2 + y) + str(r), str(y), bold=True, color=WHITE, fill=NAVY, align="right", border=True, fmt="@")
    st = {"r": hdr + 1}

    def prow(label, first_expr, growth_expr, fmt=CUR, bold=False, fill=None):
        rr = st["r"]
        cell(pr, f"A{rr}", label, bold=bold, fill=fill, size=9)
        for y in range(0, 31):
            col = get_column_letter(2 + y)
            v = first_expr(col, rr) if y == 0 else growth_expr(col, get_column_letter(2 + y - 1), rr)
            cell(pr, f"{col}{rr}", v, fmt=fmt, align="right", bold=bold, fill=fill, size=9)
        st["r"] = rr + 1
        return rr

    R_rent = prow("Rent / unit / mo",
                  lambda c, rr: f"={A['rent_unit']}",
                  lambda c, p, rr: f"={p}{rr}*(1+{A['rent_growth']})")
    R_gsr = prow("Gross scheduled rent",
                 lambda c, rr: f"={c}{R_rent}*{A['units']}*12",
                 lambda c, p, rr: f"={c}{R_rent}*{A['units']}*12")
    R_egi = prow("Effective gross income",
                 lambda c, rr: f"={c}{R_gsr}*(1-{A['vacancy_pct']})",
                 lambda c, p, rr: f"={c}{R_gsr}*(1-{A['vacancy_pct']})")
    R_opx = prow("Operating expenses",
                 lambda c, rr: f"=-{BH['opxtot']}",
                 lambda c, p, rr: f"={p}{rr}*(1+{A['exp_infl']})")
    R_noi = prow("Net operating income",
                 lambda c, rr: f"={c}{R_egi}-{c}{R_opx}",
                 lambda c, p, rr: f"={c}{R_egi}-{c}{R_opx}", bold=True)
    R_ds = prow("Debt service",
                lambda c, rr: f"={BH['ads']}",
                lambda c, p, rr: f"={BH['ads']}")
    R_cfbt = prow("Cash flow before tax",
                  lambda c, rr: f"={c}{R_noi}-{c}{R_ds}",
                  lambda c, p, rr: f"={c}{R_noi}-{c}{R_ds}", bold=True, fill=LGREY)
    bal_expr = (f"*(1+{A['refi_rate']}/12)^12-{BH['pmt']}*(((1+{A['refi_rate']}/12)^12-1)/({A['refi_rate']}/12))")
    R_bal = prow("Loan balance (EOY)",
                 lambda c, rr: f"=MAX(0,{BH['refiloan']}{bal_expr})",
                 lambda c, p, rr: f"=MAX(0,{p}{rr}{bal_expr})")
    R_val = prow("Property value",
                 lambda c, rr: f"={BH['val']}*(1+{A['apprec']})",
                 lambda c, p, rr: f"={p}{rr}*(1+{A['apprec']})")
    R_eq = prow("Equity (value − loan)",
                lambda c, rr: f"={c}{R_val}-{c}{R_bal}",
                lambda c, p, rr: f"={c}{R_val}-{c}{R_bal}", bold=True, fill="D6E4D0")
    R_ecf = prow("Investor equity cash flow",
                 lambda c, rr: f"=-{A['equity_constr']}+{BH['cashout']}",
                 lambda c, p, rr: f"={c}{R_cfbt}")
    term_col = get_column_letter(32)
    pr[f"{term_col}{R_ecf}"] = (f"={term_col}{R_cfbt}+({term_col}{R_noi}*(1+{A['rent_growth']})/{A['term_cap']})"
                                f"*(1-{A['comm_pct']}-{A['sale_close_pct']})-{term_col}{R_bal}")
    st["r"] += 1
    section(pr, st["r"], "RETURN SUMMARY (LONG HOLD)", span=6, fill=TEAL); st["r"] += 1
    sums = {}
    for label, val, fmtv in [
        ("Levered IRR", f"=IRR(B{R_ecf}:{term_col}{R_ecf})", PCT),
        ("Equity multiple", f"=SUM(C{R_ecf}:{term_col}{R_ecf})/(-B{R_ecf})", MULT),
        ("NPV @ discount rate", f"=B{R_ecf}+NPV({A['disc_rate']},C{R_ecf}:{term_col}{R_ecf})", CUR),
        ("Final-yr sale (gross)", f"={term_col}{R_noi}*(1+{A['rent_growth']})/{A['term_cap']}", CUR),
        ("Total investor cash", f"=SUM(C{R_ecf}:{term_col}{R_ecf})", CUR)]:
        rr = st["r"]
        cell(pr, f"A{rr}", label, bold=True)
        cell(pr, f"B{rr}", val, fmt=fmtv, align="right", bold=True, fill=YEL)
        sums[label] = rr
        st["r"] = rr + 1
    pr.freeze_panes = "B4"
    PRJ = {"irr": f"'30-Year Projection'!$B${sums['Levered IRR']}",
           "em": f"'30-Year Projection'!$B${sums['Equity multiple']}",
           "ecf": R_ecf, "eq": R_eq}

    # ---------------- CASH FLOW STATEMENT ----------------
    cf = wb.create_sheet("Cash Flow Statement")
    cf.sheet_view.showGridLines = False
    for col, w in zip("ABC", [42, 18, 40]):
        cf.column_dimensions[col].width = w
    r = title_band(cf, 1, "CASH FLOW STATEMENT", "Construction → refinance → operations", span=3)
    r += 1
    st = {"r": r}

    def cfrow(label, val, note="", bold=False, fill=None):
        rr = st["r"]
        cell(cf, f"A{rr}", label, bold=bold, fill=fill)
        cell(cf, f"B{rr}", val, color=GREEN, fmt=CUR, align="right", bold=bold, fill=fill)
        if note: cell(cf, f"C{rr}", note, italic=True, size=8, color="6A7480")
        st["r"] = rr + 1

    section(cf, st["r"], "CONSTRUCTION PHASE (Yr 0)", span=3, fill=NAVY); st["r"] += 1
    cfrow("Equity contribution (outflow)", f"=-{A['equity_constr']}", "includes financing costs via TPC")
    cfrow("Construction loan draws", f"={A['c_loan']}")
    cfrow("Total project cost (outflow)", f"=-{A['tpc']}")
    section(cf, st["r"], "STABILIZATION — DSCR REFINANCE", span=3, fill=NAVY); st["r"] += 1
    cfrow("DSCR refi proceeds", f"={BH['refiloan']}")
    cfrow("Construction loan payoff", f"=-{A['c_loan']}")
    cfrow("Net cash at refinance", f"={BH['cashout']}", "negative = cash-in required", bold=True, fill=YEL)
    section(cf, st["r"], "STABILIZED OPERATIONS (Yr 1)", span=3, fill=NAVY); st["r"] += 1
    cfrow("Net operating income", f"={BH['noi']}")
    cfrow("Debt service", f"=-{BH['ads']}")
    cfrow("Cash flow before tax", f"={BH['cfbt']}", "", bold=True, fill="D6E4D0")
    cf.freeze_panes = "A4"

    # ---------------- COMPARISON ----------------
    cp = wb.create_sheet("Comparison")
    cp.sheet_view.showGridLines = False
    for col, w in zip("ABCD", [34, 22, 22, 30]):
        cp.column_dimensions[col].width = w
    r = title_band(cp, 1, "BUILD-TO-SELL  vs  BUILD-TO-HOLD", "Same project, two exits", span=4)
    r += 1
    cell(cp, f"A{r}", "Metric", bold=True, color=WHITE, fill=NAVY, border=True)
    cell(cp, f"B{r}", "Build-to-Sell", bold=True, color=WHITE, fill=CLAY, align="center", border=True)
    cell(cp, f"C{r}", "Build-to-Hold", bold=True, color=WHITE, fill=TEAL, align="center", border=True)
    cell(cp, f"D{r}", "Read", bold=True, color=WHITE, fill=NAVY, border=True)
    st = {"r": r + 1}

    def cprow(label, sell, hold, read, fmt=CUR, pct=False, mult=False):
        rr = st["r"]
        f = PCT if pct else (MULT if mult else fmt)
        cell(cp, f"A{rr}", label, border=True)
        cell(cp, f"B{rr}", sell, color=GREEN, fmt=(f if sell else None), align="right", border=True)
        cell(cp, f"C{rr}", hold, color=GREEN, fmt=(f if hold else None), align="right", border=True)
        cell(cp, f"D{rr}", read, italic=True, size=9, color="4A5560", border=True, wrap=True)
        st["r"] = rr + 1

    cprow("Total profit / net cash", f"={BS['np']}",
          f"=SUM('30-Year Projection'!B{PRJ['ecf']}:AF{PRJ['ecf']})",
          "Sell = one-time; Hold = net long-hold cash")
    cprow("Cash required (equity)", f"={BS['eq']}", f"={BH['eqrem']}", "Hold equity after refi cash movement")
    cprow("IRR", f"={BS['irr']}", f"={PRJ['irr']}", "Annualized", pct=True)
    cprow("Return on cost / CoC", f"={BS['roc']}", f"={BH['coc']}", "", pct=True)
    cprow("Equity multiple", f"=({BS['eq']}+{BS['np']})/{BS['eq']}", f"={PRJ['em']}", "", mult=True)
    cprow("Year-1 DSCR", "", f"={BH['dscr']}", "lender test ≥1.20–1.25", mult=True)
    cprow("Liquidity / tax", "High / ordinary gain", "Low / depreciation + 1031", "", fmt=None)
    cp.freeze_panes = "A4"

    # ---------------- SENSITIVITY ----------------
    se = wb.create_sheet("Sensitivity Analysis")
    se.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFG", [30, 15, 15, 15, 15, 15, 15]):
        se.column_dimensions[col].width = w
    r = title_band(se, 1, "SENSITIVITY ANALYSIS", "Sell net profit & Hold Yr-1 CFBT", span=7)
    r += 1
    section(se, r, "SELL NET PROFIT — Sale $/unit (rows) × Cost variance (cols)", span=7, fill=NAVY); r += 1
    cell(se, f"A{r}", "Sale $/unit \\ Cost", bold=True, fill=LGREY, border=True)
    for i, cv in enumerate(C["sens_cost_var"]):
        cell(se, get_column_letter(2 + i) + str(r), cv, bold=True, fill=LGREY, align="center", border=True, fmt=PCT)
    r += 1
    for sv in C["sens_sale_var"]:
        cell(se, f"A{r}", f"={A['sale_unit']}*(1+{sv})", fmt=CUR, border=True, bold=True, fill=LGREY)
        for i, cv in enumerate(C["sens_cost_var"]):
            f = (f"=({A['sale_unit']}*(1+{sv})*{A['units']})*(1-{A['comm_pct']}-{A['sale_close_pct']})"
                 f"-{A['tpc']}*(1+{cv})")
            cell(se, get_column_letter(2 + i) + str(r), f, fmt=CUR, align="right", border=True)
        r += 1
    r += 1
    section(se, r, "HOLD Yr-1 CFBT — Rent/unit (rows) × Cap rate (cols)", span=7, fill=NAVY); r += 1
    cell(se, f"A{r}", "Rent/unit \\ Cap", bold=True, fill=LGREY, border=True)
    for i, cpv in enumerate(C["sens_caps"]):
        cell(se, get_column_letter(2 + i) + str(r), cpv, bold=True, fill=LGREY, align="center", border=True, fmt=PCT)
    r += 1
    for rv in C["sens_rents"]:
        cell(se, f"A{r}", rv, fmt=CUR, border=True, bold=True, fill=LGREY)
        for i, cpv in enumerate(C["sens_caps"]):
            noi = (f"({rv}*{A['units']}*12*(1-{A['vacancy_pct']})*(1-{A['mgmt_pct']}-{A['rm_pct']}-{A['lease_pct']})"
                   f"-{A['ins_unit']}*{A['units']}-{A['reserve_unit']}*{A['units']}-{A['util_admin']}"
                   f"-{A['tax_pct_val']}*{A['tpc']})")
            f = f"={noi}-(-PMT({A['refi_rate']}/12,{A['amort_yrs']}*12,{A['refi_ltv']}*{noi}/{cpv}))"
            cell(se, get_column_letter(2 + i) + str(r), f, fmt=CUR, align="right", border=True)
        r += 1
    se.freeze_panes = "A4"

    # ---------------- MARKET SUMMARY ----------------
    ms = wb.create_sheet("Market Summary")
    ms.sheet_view.showGridLines = False
    for col, w in zip("ABCD", [34, 26, 40, 2]):
        ms.column_dimensions[col].width = w
    r = title_band(ms, 1, "MARKET SUMMARY", "Fill from market research; tag [REFRESHED]/[FILE]/[ESTIMATE]", span=3)
    r += 1
    for sec_name, rows in C["market_rows"]:
        section(ms, r, sec_name, span=3, fill=NAVY); r += 1
        for metric, value, src in rows:
            cell(ms, f"A{r}", metric, border=True)
            cell(ms, f"B{r}", value, align="right", border=True)
            cell(ms, f"C{r}", src, italic=True, size=8, color="6A7480", border=True)
            r += 1
    ms.freeze_panes = "A4"

    # ---------------- EXECUTIVE DASHBOARD ----------------
    xd = wb.create_sheet("Executive Dashboard")
    xd.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", [30, 18, 4, 30, 18, 4]):
        xd.column_dimensions[col].width = w
    r = title_band(xd, 1, "EXECUTIVE DASHBOARD", C["subtitle"], span=6)
    r += 1

    def kpi(col, row, label, ref, fmt=CUR):
        cell(xd, f"{col}{row}", label, bold=True, size=9, color="4A5560", fill=LGREY, align="left")
        cell(xd, f"{col}{row + 1}", ref, bold=True, size=14, color=TEAL, fmt=fmt, align="left", fill=LGREY)

    section(xd, r, "PROJECT SNAPSHOT", span=6, fill=NAVY); rr = r + 1
    kpi("A", rr, "Units", f"={A['units']}", fmt=NUM); kpi("D", rr, "Total Project Cost", f"={A['tpc']}")
    rr += 2
    kpi("A", rr, "Cost / unit", f"={A['cost_unit']}"); kpi("D", rr, "Equity required", f"={A['equity_constr']}")
    r = rr + 2
    section(xd, r, "BUILD-TO-SELL", span=6, fill=CLAY); rr = r + 1
    kpi("A", rr, "Net profit", f"={BS['np']}"); kpi("D", rr, "Annualized IRR", f"={BS['irr']}", fmt=PCT)
    r = rr + 2
    section(xd, r, "BUILD-TO-HOLD", span=6, fill=TEAL); rr = r + 1
    kpi("A", rr, "Year-1 NOI", f"={BH['noi']}"); kpi("D", rr, "Stabilized value", f"={BH['val']}")
    rr += 2
    kpi("A", rr, "Year-1 DSCR", f"={BH['dscr']}", fmt=MULT); kpi("D", rr, "Long-hold IRR", f"={PRJ['irr']}", fmt=PCT)
    r = rr + 2
    section(xd, r, "RISK INDICATORS", span=6, fill=NAVY); r += 1
    cell(xd, f"A{r}", "Indicator", bold=True, fill=NAVY, color=WHITE, border=True)
    cell(xd, f"B{r}", "Status", bold=True, fill=NAVY, color=WHITE, border=True, align="center")
    cell(xd, f"D{r}", "Basis", bold=True, fill=NAVY, color=WHITE, border=True); r += 1
    cell(xd, f"A{r}", "Financing (DSCR ≥1.20)", border=True)
    cell(xd, f"B{r}", f'=IF({BH["dscr"]}>=1.25,"STRONG",IF({BH["dscr"]}>=1.10,"ADEQUATE","TIGHT"))', border=True, align="center", bold=True)
    cell(xd, f"D{r}", f'=TEXT({BH["dscr"]},"0.00")&"x DSCR"', border=True, color="6A7480", italic=True); r += 1
    cell(xd, f"A{r}", "Sell profitability (margin ≥15%)", border=True)
    cell(xd, f"B{r}", f'=IF({BS["np"]}/{A["gross_sales"]}>=0.15,"STRONG",IF({BS["np"]}/{A["gross_sales"]}>=0.08,"THIN","NEGATIVE"))', border=True, align="center", bold=True)
    cell(xd, f"D{r}", f'=TEXT({BS["np"]}/{A["gross_sales"]},"0.0%")&" margin"', border=True, color="6A7480", italic=True); r += 1
    cell(xd, f"A{r}", "Hold viability (yield on cost vs cap)", border=True)
    cell(xd, f"B{r}", f'=IF({BH["capr"]}>={A["exit_cap"]},"CREATES VALUE","VALUE DESTRUCTIVE")', border=True, align="center", bold=True)
    cell(xd, f"D{r}", f'=TEXT({BH["capr"]},"0.0%")&" vs "&TEXT({A["exit_cap"]},"0.0%")&" cap"', border=True, color="6A7480", italic=True); r += 1
    r += 1
    section(xd, r, "RECOMMENDATION", span=6, fill=TEAL); r += 1
    xd.merge_cells(start_row=r, start_column=1, end_row=r + 3, end_column=6)
    cell(xd, f"A{r}", C["recommendation"], wrap=True, valign="top", size=10, fill=PAPER)
    xd.row_dimensions[r].height = 90
    xd.freeze_panes = "A3"

    # ---------------- CHARTS ----------------
    ch = wb.create_sheet("Charts")
    ch.sheet_view.showGridLines = False
    r = title_band(ch, 1, "CHARTS", "Auto-linked to model outputs", span=8)
    cell(ch, "A4", "Cost component", bold=True); cell(ch, "B4", "Amount", bold=True)
    comps = [("Land", A['land']), ("Hard construction", A['hard_total']),
             ("Soft + fees", f"{A['plans_design']}+{A['bldg_permit']}+{A['impact_total']}+{A['contingency']}+{A['tax_constr']}+{A['ins_constr']}+{A['legal_misc']}"),
             ("Financing", A['c_fin_cost'])]
    rr = 5
    for name, ref in comps:
        cell(ch, f"A{rr}", name); cell(ch, f"B{rr}", f"={ref}", color=GREEN, fmt=CUR); rr += 1
    bar = BarChart(); bar.title = "Total Project Cost Breakdown"; bar.type = "col"; bar.style = 10
    data = Reference(ch, min_col=2, min_row=4, max_row=rr - 1)
    cats = Reference(ch, min_col=1, min_row=5, max_row=rr - 1)
    bar.add_data(data, titles_from_data=True); bar.set_categories(cats); bar.height = 8; bar.width = 15
    ch.add_chart(bar, "D4")
    line = LineChart(); line.title = "Long-Hold Equity Growth"; line.style = 12
    edata = Reference(pr, min_col=2, min_row=PRJ['eq'], max_col=32, max_row=PRJ['eq'])
    line.add_data(edata, from_rows=True, titles_from_data=False)
    line.height = 8; line.width = 20
    ch.add_chart(line, "D22")
    for col in "ABC":
        ch.column_dimensions[col].width = 22

    order = ["Executive Dashboard", "Build-to-Sell", "Build-to-Hold", "Comparison",
             "30-Year Projection", "Construction Loan Model", "P&L Year 1", "Cash Flow Statement",
             "Development Budget", "Market Summary", "Assumptions", "Sensitivity Analysis", "Charts"]
    for i, nm in enumerate(order):
        wb.move_sheet(nm, -(wb.sheetnames.index(nm)) + i)
    wb.save(out_path)
    return out_path


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--config", help="JSON file overriding DEFAULT_CONFIG keys")
    ap.add_argument("--out", default="buy-vs-build-model.xlsx", help="output xlsx path")
    args = ap.parse_args()
    C = dict(DEFAULT_CONFIG)
    if args.config:
        with open(args.config) as f:
            C.update(json.load(f))
    path = build(C, args.out)
    print(f"Saved: {path}")
    print("NEXT: python3 validate_model.py", path)


if __name__ == "__main__":
    main()

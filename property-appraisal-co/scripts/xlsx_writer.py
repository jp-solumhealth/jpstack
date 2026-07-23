"""Write the appraisal XLSX with Summary, Sale comps, Rent comps, Enrichment, Log tabs."""
from __future__ import annotations

from dataclasses import asdict
from pathlib import Path
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .scrapers.base import Listing, Subject

DISCLAIMER = (
    "Indicative valuation only — not a certified avalúo (Resolución SNR / RNA). "
    "Comps reflect listing prices, not closed sales. Apply a market-specific "
    "adjustment if used for underwriting."
)

HEADER_FILL = PatternFill("solid", fgColor="0F2A4A")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _fmt_cop(v: float | None) -> str:
    if v is None:
        return "—"
    return f"${v:,.0f}"


def _h(ws, row: int, label: str, value: Any) -> None:
    ws.cell(row=row, column=1, value=label).font = Font(bold=True)
    ws.cell(row=row, column=2, value=value)


def _header_row(ws, row: int, headers: list[str]) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")


def _autosize(ws, max_col: int) -> None:
    for col_idx in range(1, max_col + 1):
        max_len = 10
        for cell in ws[get_column_letter(col_idx)]:
            v = cell.value
            if v is None:
                continue
            max_len = max(max_len, min(60, len(str(v))))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2


def write(
    *,
    subject: Subject,
    valuation: dict,
    rent: dict,
    enrichment: dict,
    loan: dict,
    sale_comps: list[Listing],
    rent_comps: list[Listing],
    log: list[dict],
    out_path: str | Path,
) -> Path:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # ----- Summary -----
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Property Appraisal — Indicative"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:D1")

    ws["A3"] = "Generated"
    ws["B3"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")

    row = 5
    _h(ws, row, "Subject", ""); row += 1
    for k in ["city", "barrio", "address", "m2_built", "m2_private", "bedrooms",
              "bathrooms", "parking", "estrato", "year_built", "admin_fee",
              "matricula", "building_name", "listing_url"]:
        _h(ws, row, k, getattr(subject, k, None)); row += 1
    row += 1

    appr = valuation.get("appraised", {})
    pps = valuation.get("price_per_m2", {})
    _h(ws, row, "Valuation", ""); row += 1
    _h(ws, row, "n comps", valuation.get("n_comps", 0)); row += 1
    _h(ws, row, "Confidence", valuation.get("confidence", "low")); row += 1
    _h(ws, row, "Price/m² P25", _fmt_cop(pps.get("p25"))); row += 1
    _h(ws, row, "Price/m² P50", _fmt_cop(pps.get("p50"))); row += 1
    _h(ws, row, "Price/m² P75", _fmt_cop(pps.get("p75"))); row += 1
    _h(ws, row, "Appraised LOW", _fmt_cop(appr.get("low"))); row += 1
    _h(ws, row, "Appraised MID", _fmt_cop(appr.get("mid"))); row += 1
    _h(ws, row, "Appraised HIGH", _fmt_cop(appr.get("high"))); row += 1
    row += 1

    rent_p = rent.get("rent", {})
    _h(ws, row, "Rent estimate", ""); row += 1
    _h(ws, row, "n rent comps", rent.get("n_rent_comps", 0)); row += 1
    _h(ws, row, "Rent P25 / mo", _fmt_cop(rent_p.get("p25"))); row += 1
    _h(ws, row, "Rent P50 / mo", _fmt_cop(rent_p.get("p50"))); row += 1
    _h(ws, row, "Rent P75 / mo", _fmt_cop(rent_p.get("p75"))); row += 1
    gy = rent.get("gross_yield_pct")
    _h(ws, row, "Gross yield (%)", f"{gy:.2f}%" if gy is not None else "—"); row += 1
    row += 1

    _h(ws, row, "Loan sizing", ""); row += 1
    _h(ws, row, "LTV", f"{loan.get('ltv', 0):.0%}"); row += 1
    _h(ws, row, "Suggested max loan (COP)", _fmt_cop(loan.get("max_loan_cop"))); row += 1
    row += 1

    _h(ws, row, "Enrichment", ""); row += 1
    _h(ws, row, "Constructora", enrichment.get("constructora_final")); row += 1
    _h(ws, row, "Año construcción", enrichment.get("anio_final")); row += 1
    _h(ws, row, "Proyecto", enrichment.get("proyecto")); row += 1
    cb = enrichment.get("catastro_bogota") or {}
    cm = enrichment.get("catastro_medellin") or {}
    _h(ws, row, "Catastro Bogotá", cb.get("status", "—")); row += 1
    _h(ws, row, "Catastro Medellín", cm.get("status", "—")); row += 1
    snr = enrichment.get("snr") or {}
    _h(ws, row, "SNR / matrícula", snr.get("status", "—")); row += 1
    row += 2

    ws.cell(row=row, column=1, value="Disclaimer").font = Font(bold=True, italic=True)
    ws.cell(row=row, column=2, value=DISCLAIMER).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[row].height = 40
    ws.column_dimensions["B"].width = 80

    _autosize(ws, max_col=4)

    # ----- Sale comps -----
    _comp_sheet(wb, "Sale comps", sale_comps)
    _comp_sheet(wb, "Rent comps", rent_comps)

    # ----- Enrichment -----
    we = wb.create_sheet("Enrichment")
    we["A1"] = "Field"; we["B1"] = "Value"
    we["A1"].font = Font(bold=True); we["B1"].font = Font(bold=True)
    r = 2
    for k, v in _flatten(enrichment).items():
        we.cell(row=r, column=1, value=k)
        we.cell(row=r, column=2, value=str(v) if v is not None else "—")
        r += 1
    _autosize(we, max_col=2)

    # ----- Log -----
    wl = wb.create_sheet("Log")
    headers = ["portal", "transaction", "returned", "kept", "dropped", "notes"]
    _header_row(wl, 1, headers)
    for i, entry in enumerate(log, start=2):
        for j, h in enumerate(headers, start=1):
            wl.cell(row=i, column=j, value=entry.get(h))
    _autosize(wl, max_col=len(headers))

    wb.save(out_path)
    return out_path


def _comp_sheet(wb: Workbook, title: str, comps: list[Listing]) -> None:
    ws = wb.create_sheet(title)
    headers = ["source", "url", "title", "price_cop", "m2_built", "price_per_m2",
               "bedrooms", "bathrooms", "parking", "estrato", "barrio", "city",
               "listing_date", "year_built", "constructora", "proyecto"]
    _header_row(ws, 1, headers)
    for i, c in enumerate(comps, start=2):
        d = asdict(c)
        d["price_per_m2"] = c.price_per_m2
        for j, h in enumerate(headers, start=1):
            ws.cell(row=i, column=j, value=d.get(h))
    _autosize(ws, max_col=len(headers))


def _flatten(d: dict, prefix: str = "") -> dict:
    out: dict[str, Any] = {}
    for k, v in d.items():
        key = f"{prefix}.{k}" if prefix else k
        if isinstance(v, dict):
            out.update(_flatten(v, key))
        else:
            out[key] = v
    return out
